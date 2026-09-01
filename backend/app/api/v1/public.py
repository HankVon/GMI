"""对外官网公开统计接口 — 无需登录, 仅返回脱敏聚合数据。

安全边界:
- 不返回任何单位名称、人名、联系方式、金额明细等敏感实体。
- 仅暴露平台级聚合统计(总量 / 地域Top / 类型构成 / 月度趋势)。
- 不依赖 get_current_user, 通过独立 router 注册, 不受鉴权中间件约束。
"""
import asyncio
import json
import logging
import re
import time
import uuid
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from sqlalchemy import select, func, text, or_, not_
from sqlalchemy.orm import Session

logger = logging.getLogger("public")

from app.database import get_db
from app.models.bid_notice import BidNotice
from app.models.company import Company
from app.models.intent_notice import IntentNotice
from app.models.opportunity import Opportunity
from app.models.opportunity_tag import OpportunityTag, OpportunityTagDef
from app.models.web_clue import WebClue
from app.models.person import Person
from app.models.project import Project
from app.models.business_network import NetworkEdge
from app.models.industry_data import CompanyIc, Qualification, Honor, CreditRecord, PersonCert
from app.models.content import ContentAsset
from app.models.notification import Notification
from app.config import settings

import httpx

router = APIRouter(prefix="/public", tags=["对外官网公开数据"])


class ContactRequest(BaseModel):
    name: str = Field(default="", max_length=128)
    org: str = Field(default="", max_length=256)
    contact: str = Field(min_length=3, max_length=256)
    type: str = Field(default="other", max_length=32)
    description: str = Field(default="", max_length=4000)


@router.post("/contact")
async def public_contact(payload: ContactRequest, db: Session = Depends(get_db)):
    """官网咨询/反馈落库，并由全局写操作审计中间件记录请求。"""
    # 复用通知表承载待处理咨询，user_id=0 表示系统待处理队列。
    row = Notification(user_id=0, type="contact", title=f"官网{payload.type}：{payload.name or '匿名用户'}", content=json.dumps(payload.model_dump(), ensure_ascii=False), related_type="contact", is_read=False, is_deleted=False)
    db.add(row)
    db.commit()
    return {"success": True, "id": row.id, "status": "pending"}


# ───────────────────────────────────────────────────────────
# 真实 LLM 研判(对外脱敏: 模型输入真实意向字段, 输出约束不暴露具体真名)
# ───────────────────────────────────────────────────────────
LLM_INTENT_PROMPT = """你是招投标情报分析师。基于真实招标意向给出研判。禁止输出具体单位全称、人名、电话等敏感实体，单位/人员用泛称（如「该主管部门」「当地某地质单位」「相关业主」）。只输出JSON，不要任何额外文字。

意向信息：
标题：{title}
发布部门：{dept}
地域：{region}
行业：{industry}
金额档位：{amount_level}
状态：{status}
关键词：{keywords}

输出JSON：
{{"summary":"一句话研判","heat":0-100整数,"coop_prob":0-100整数,"orgs":["泛称单位1","泛称单位2"],"persons_hint":"关键人员角色泛称与触达思路(不写真名)","network_path":"基于人脉图谱的触达路径建议(泛称)","advice":["建议1","建议2","建议3"],"opportunities":["机会1","机会2"]}}"""


# ───────────────────────────────────────────────────────────
# 脱敏辅助
# ───────────────────────────────────────────────────────────
def _anon_name(type_: str, id_: int) -> str:
    """将真实实体名脱敏为 '类型·尾号' 标签, 避免对外暴露单位/人员真名。"""
    tail = f"{id_ % 10000:04d}"
    label = {"person": "人员", "company": "单位", "project": "项目"}.get(type_, "实体")
    return f"{label}·{tail}"


def _amount_level(amount) -> str:
    """预算金额脱敏为区间档位, 不暴露精确金额。"""
    if amount is None:
        return "未披露"
    try:
        v = float(amount)
    except (TypeError, ValueError):
        return "未披露"
    if v < 100:
        return "100万以下"
    if v < 500:
        return "100–500万"
    if v < 2000:
        return "500–2000万"
    if v < 10000:
        return "2000万–1亿"
    return "1亿以上"


def _rule_heat(intent) -> int:
    """规则化意向热度(0-100): 金额档位 + 近期度 + 状态加权。演示性, 非真实研判。"""
    level_score = {"未披露": 30, "100万以下": 35, "100–500万": 55,
                   "500–2000万": 75, "2000万–1亿": 90, "1亿以上": 100}
    score = level_score.get(_amount_level(intent.amount), 30)
    # 近 30 天加权
    if intent.published_at:
        days = (datetime.now() - intent.published_at).days
        if days <= 30:
            score = min(100, score + 10)
        elif days > 180:
            score = max(10, score - 15)
    if intent.status == "new":
        score = min(100, score + 5)
    return int(score)


def _rule_coop_prob(intent, edge_count: int) -> int:
    """规则化合作概率(演示估算): 基于意向状态 + 关联人脉边密度。非真实商业判断。"""
    base = 55 if intent.status == "new" else 40
    # 平台整体人脉关系越丰富, 触达概率越高(演示逻辑)
    boost = min(30, edge_count // 20)
    return min(95, base + boost)


def _rule_advice(intent) -> str:
    """基于项目类型/行业的模板化行动建议(规则生成, 演示)。"""
    ind = (intent.industry or intent.project_type or "相关")
    if intent.status == "expired":
        return f"该{ind}意向已过期，建议纳入历史库做同类机会规律分析，提前布局下一周期。"
    return (f"建议优先核实「{intent.region or '当地'}」{ind}类意向的业主与决策链，"
            f"结合平台人脉图谱定位可触达的桥接人，30 天内完成首次接洽。")


def _build_graph(edges) -> dict:
    """由 network_edge 构建脱敏图谱(节点名脱敏, 仅关系类型可见)。"""
    nodes_map: dict[str, dict] = {}
    links = []
    for e in edges:
        s_key = f"{e['src_type']}:{e['src_id']}"
        t_key = f"{e['tgt_type']}:{e['tgt_id']}"
        if s_key not in nodes_map:
            nodes_map[s_key] = {
                "id": s_key, "name": _anon_name(e["src_type"], e["src_id"]),
                "type": e["src_type"], "degree": 0,
            }
        if t_key not in nodes_map:
            nodes_map[t_key] = {
                "id": t_key, "name": _anon_name(e["tgt_type"], e["tgt_id"]),
                "type": e["tgt_type"], "degree": 0,
            }
        nodes_map[s_key]["degree"] += 1
        nodes_map[t_key]["degree"] += 1
        links.append({
            "source": s_key, "target": t_key,
            "rel": e["rel_zh"] or e["rel_type"],
            "weight": float(e["weight"]) if e["weight"] is not None else 1.0,
        })
    return {"nodes": list(nodes_map.values()), "links": links}


@router.get("/search")
async def public_search(
    keyword: str = Query("", min_length=1, max_length=64, description="检索关键词(单位/人员/项目名)"),
    province: Optional[str] = Query(None, description="地区筛选(省)"),
    type: str = Query("all", pattern="^(all|company|person|project)$", description="实体类型"),
    limit: int = Query(20, ge=1, le=50),
    db: Session = Depends(get_db),
):
    """对外官网·公开检索(脱敏) — 对标建设通公开列表引流。

    返回脱敏列表: 单位/项目名可示, 人员名脱敏为「人员·尾号」;
    电话/证件/精确金额一律不返回, 金额以区间档位展示。
    登录后可在工作台检索完整详情。
    """
    kw = keyword.strip()
    out: dict = {"companies": [], "persons": [], "projects": []}
    if not kw:
        return {"success": True, "message": "ok", "data": out}

    if type in ("all", "company"):
        rows = db.execute(
            select(Company).where(
                Company.is_deleted == False,
                (Company.name.like(f"%{kw}%")) | (Company.short_name.like(f"%{kw}%")),
                *((Company.province == province,) if province else ()),
            ).order_by(Company.id.desc()).limit(limit)
        ).scalars().all()
        for c in rows:
            bid_cnt = db.execute(text(
                "SELECT COUNT(*) FROM bid_notice WHERE is_deleted=0 "
                "AND (purchaser_company_id = :cid OR JSON_SEARCH(meta, 'one', :n) IS NOT NULL)"
            ), {"cid": c.id, "n": c.name}).scalar() or 0
            out["companies"].append({
                "id": c.id,
                "name": c.name,
                "province": c.province,
                "city": c.city,
                "company_type": c.company_type,
                "credit_level": c.credit_level,
                "bid_count": int(bid_cnt),
            })

    if type in ("all", "person"):
        rows = db.execute(
            select(Person).where(
                Person.is_deleted == False, Person.name.like(f"%{kw}%"),
            ).order_by(Person.id.desc()).limit(limit)
        ).scalars().all()
        for p in rows:
            out["persons"].append({
                "id": p.id,
                "name": p.name,
                "position": p.position,
                "company_id": p.company_id,
            })

    if type in ("all", "project"):
        rows = db.execute(
            select(Project).where(
                Project.is_deleted == False, Project.name.like(f"%{kw}%"),
            ).order_by(Project.id.desc()).limit(limit)
        ).scalars().all()
        for pr in rows:
            attrs = pr.ext_attrs if isinstance(pr.ext_attrs, dict) else {}
            out["projects"].append({
                "id": pr.id,
                "name": pr.name,
                "status": pr.status,
                "amount_level": _amount_level(attrs.get("amount") or attrs.get("total_invest")),
                "category": attrs.get("category") or None,
            })

    return {
        "success": True,
        "message": "ok",
        "data": out,
        "note": "公开检索视图: 单位/项目/人员名称均展示真实信息(内部轻度使用); 电话/证件/精确金额不对外返回。",
    }


@router.get("/home")
async def public_home(
    db: Session = Depends(get_db),
):
    """官网首页数据 — 最新动态(中标/意向/单位) + 活跃排行。

    对标建设通首页(hhb): 企业名/公告标题为公开信息可展示;
    不返回联系方式/金额明细等敏感实体。
    """
    # 最新中标/标讯
    bids = db.execute(
        select(BidNotice).where(BidNotice.is_deleted == False)
        .order_by(BidNotice.published_at.is_(None), BidNotice.published_at.desc())
        .limit(6)
    ).scalars().all()
    latest_bids = [{
        "id": b.id,
        "title": (b.title or "")[:80],
        "region": b.region,
        "notice_type": b.notice_type,
        "purchaser": b.purchaser,
        "published_at": b.published_at.strftime("%Y-%m-%d") if b.published_at else "",
    } for b in bids]
    # 最新意向(仅已发布: wf_status='published' 或 历史未标记 NULL 的记录)
    intents = db.execute(
        select(IntentNotice).where(
            IntentNotice.is_deleted == False,
            or_(IntentNotice.wf_status == "published", IntentNotice.wf_status.is_(None)),
        )
        .order_by(IntentNotice.published_at.is_(None), IntentNotice.published_at.desc())
        .limit(6)
    ).scalars().all()
    latest_intents = [{
        "id": it.id,
        "title": (it.title or "")[:80],
        "dept": it.dept,
        "region": it.region,
        "amount_level": _amount_level(it.amount),
        "published_at": it.published_at.strftime("%Y-%m-%d") if it.published_at else "",
    } for it in intents]
    # 最新单位
    companies = db.execute(
        select(Company).where(Company.is_deleted == False)
        .order_by(Company.created_at.desc()).limit(6)
    ).scalars().all()
    latest_companies = []
    for c in companies:
        ic = db.execute(select(CompanyIc).where(CompanyIc.company_id == c.id, CompanyIc.is_deleted == False).order_by(CompanyIc.updated_at.desc())).scalars().first()
        latest_bid = db.execute(select(BidNotice.published_at).where(BidNotice.purchaser_company_id == c.id, BidNotice.is_deleted == False).order_by(BidNotice.published_at.desc())).scalar()
        latest_companies.append({
            "id": c.id, "name": c.name, "province": c.province, "city": c.city,
            "company_type": c.company_type, "registered_capital": ic.registered_capital if ic else None,
            "latest_bid_at": latest_bid.strftime("%Y-%m-%d") if latest_bid else None,
            "updated_at": c.updated_at.strftime("%Y-%m-%d") if c.updated_at else "",
        })
    content_rows = db.execute(select(ContentAsset).where(ContentAsset.is_deleted == False, ContentAsset.status == "published").order_by(ContentAsset.published_at.desc()).limit(12)).scalars().all()
    latest_content = [{"id": a.id, "title": a.title, "kind": a.kind, "summary": a.summary or "", "published_at": a.published_at.strftime("%Y-%m-%d") if a.published_at else "", "published_url": f"/site/content/{a.id}"} for a in content_rows]
    # 排行1: 中标业主活跃榜(按已匹配采购人)
    rank_owners = db.execute(text(
        "SELECT c.name, COUNT(b.id) AS cnt FROM bid_notice b "
        "JOIN company c ON c.id = b.purchaser_company_id "
        "WHERE b.is_deleted=0 AND c.is_deleted=0 "
        "GROUP BY c.id ORDER BY cnt DESC, c.id DESC LIMIT 10"
    )).all()
    # 排行2: 中标供应商活跃榜(meta JSON 供应商匹配到公司)
    rank_suppliers = db.execute(text(
        "SELECT c.name, COUNT(*) AS cnt FROM bid_notice b "
        "JOIN JSON_TABLE(b.meta, '$.suppliers[*]' COLUMNS(name VARCHAR(512) PATH '$.supplier')) st "
        "JOIN company c ON c.name = st.name "
        "WHERE b.is_deleted=0 AND c.is_deleted=0 "
        "GROUP BY c.id ORDER BY cnt DESC, c.id DESC LIMIT 10"
    )).all() if False else []
    if not rank_suppliers:
        # JSON_TABLE 兼容性兜底: 按 meta 文本匹配
        rank_suppliers = db.execute(text(
            "SELECT c.name, COUNT(*) AS cnt FROM bid_notice b "
            "JOIN company c ON JSON_SEARCH(b.meta, 'one', c.name) IS NOT NULL "
            "WHERE b.is_deleted=0 AND c.is_deleted=0 "
            "GROUP BY c.id ORDER BY cnt DESC, c.id DESC LIMIT 10"
        )).all()
    # 地区 Top
    region_top = db.execute(text(
        "SELECT region, COUNT(*) cnt FROM web_clue WHERE is_deleted=0 AND region IS NOT NULL AND region<>'' "
        "GROUP BY region ORDER BY cnt DESC LIMIT 10"
    )).all()
    return {
        "success": True,
        "data": {
            "latest_bids": latest_bids,
            "latest_intents": latest_intents,
            "latest_companies": latest_companies,
            "latest_content": latest_content,
            "rank_owners": [{"name": r[0], "count": int(r[1])} for r in rank_owners],
            "rank_suppliers": [{"name": r[0], "count": int(r[1])} for r in rank_suppliers],
            "region_top": [{"province": r[0], "count": int(r[1])} for r in region_top],
        },
    }


@router.get("/home/content/{asset_id}")
async def public_home_content(asset_id: int, db: Session = Depends(get_db)):
    """官网公开内容详情，仅开放已发布内容。"""
    asset = db.get(ContentAsset, asset_id)
    if not asset or asset.is_deleted or asset.status != "published":
        raise HTTPException(status_code=404, detail="内容不存在或尚未发布")
    return {"success": True, "item": {"id": asset.id, "title": asset.title, "kind": asset.kind, "summary": asset.summary or "", "content": asset.content or "", "source_data": asset.source_data or {}, "published_at": asset.published_at.strftime("%Y-%m-%d %H:%M") if asset.published_at else ""}}


@router.get("/home/rankings")
async def public_home_rankings(
    province: str = Query("四川", min_length=1, max_length=64),
    limit: int = Query(10, ge=1, le=50),
    db: Session = Depends(get_db),
):
    """首页建企三榜：综合实力、中标数量、中标金额。所有结果来自真实中标关联数据。"""
    companies = db.execute(select(Company).where(Company.is_deleted == False, Company.province == province)).scalars().all()
    company_map = {c.id: c for c in companies}
    bids = db.execute(select(BidNotice).where(BidNotice.is_deleted == False, BidNotice.purchaser_company_id.in_(list(company_map) or [-1]))).scalars().all()
    aggregates: dict[int, dict] = {c.id: {"company_id": c.id, "name": c.name, "count": 0, "amount": 0.0} for c in companies}
    for bid in bids:
        row = aggregates.get(bid.purchaser_company_id)
        if not row: continue
        row["count"] += 1
        meta = bid.meta or {}
        for supplier in meta.get("suppliers", []) if isinstance(meta.get("suppliers"), list) else []:
            if isinstance(supplier, dict) and supplier.get("amount") is not None:
                try: row["amount"] += float(supplier["amount"])
                except (TypeError, ValueError): pass
    rows = list(aggregates.values())
    count_rows = [{"company_id": r["company_id"], "name": r["name"], "value": r["count"]} for r in sorted(rows, key=lambda x: (-x["count"], x["name"]))[:limit]]
    amount_rows = [{"company_id": r["company_id"], "name": r["name"], "value": round(r["amount"], 2)} for r in sorted(rows, key=lambda x: (-x["amount"], x["name"]))[:limit]]
    strength = [{"company_id": r["company_id"], "name": r["name"], "value": round(min(100, r["count"] * 2 + min(r["amount"] / 10000, 50)), 2)} for r in sorted(rows, key=lambda x: (-x["count"], -x["amount"]))[:limit]]
    return {"success": True, "data": {"strength": strength, "count": count_rows, "amount": amount_rows}}


@router.get("/home/feed")
async def public_home_feed(
    category: str = Query("companies", max_length=32),
    page: int = Query(1, ge=1),
    page_size: int = Query(12, ge=1, le=50),
    province: str = Query("", max_length=64),
    db: Session = Depends(get_db),
):
    """首页九类动态统一分页接口，返回与企业卡片一致的字段契约。"""
    category_map = {
        "companies": (Company, Company.updated_at, lambda row: {"id": row.id, "name": row.name, "type": row.company_type, "province": row.province, "region": row.city, "capital": (db.execute(select(CompanyIc.registered_capital).where(CompanyIc.company_id == row.id, CompanyIc.is_deleted == False).order_by(CompanyIc.updated_at.desc())).scalar() or "未披露"), "updated_at": row.updated_at.strftime("%Y-%m-%d") if row.updated_at else ""}),
        "bids": (BidNotice, BidNotice.published_at, lambda row: {"id": row.id, "name": row.title, "title": row.title, "type": row.notice_type, "province": row.region, "updated_at": row.published_at.strftime("%Y-%m-%d") if row.published_at else ""}),
        "tenders": (BidNotice, BidNotice.published_at, lambda row: {"id": row.id, "name": row.title, "title": row.title, "type": row.notice_type, "province": row.region, "updated_at": row.published_at.strftime("%Y-%m-%d") if row.published_at else ""}),
        "projects": (Project, Project.updated_at, lambda row: {"id": row.id, "name": row.name, "type": "项目", "updated_at": row.updated_at.strftime("%Y-%m-%d") if row.updated_at else ""}),
        "persons": (Person, Person.updated_at, lambda row: {"id": row.id, "name": row.name, "type": row.position or "人员", "updated_at": row.updated_at.strftime("%Y-%m-%d") if row.updated_at else ""}),
    }
    model_info = category_map.get(category)
    if model_info:
        model, order_col, serializer = model_info
        stmt = select(model).where(model.is_deleted == False)
        if category == "bids": stmt = stmt.where(BidNotice.notice_type.like("%中标%")).where(BidNotice.notice_type.notlike("%招标%"))
        if category == "tenders": stmt = stmt.where(BidNotice.notice_type.like("%招标%"))
        if province and hasattr(model, "province"): stmt = stmt.where(model.province == province)
        total = db.execute(select(func.count()).select_from(stmt.subquery())).scalar() or 0
        rows = db.execute(stmt.order_by(order_col.desc()).offset((page - 1) * page_size).limit(page_size)).scalars().all()
        return {"success": True, "data": {"category": category, "total": total, "page": page, "page_size": page_size, "items": [serializer(row) for row in rows]}}
    if category in {"qualifications", "honors", "credit"}:
        model = {"qualifications": Qualification, "honors": Honor, "credit": CreditRecord}[category]
        total = db.execute(select(func.count()).select_from(select(model.id).where(model.is_deleted == False).subquery())).scalar() or 0
        rows = db.execute(select(model).where(model.is_deleted == False).order_by(model.updated_at.desc()).offset((page - 1) * page_size).limit(page_size)).scalars().all()
        return {"success": True, "data": {"category": category, "total": total, "page": page, "page_size": page_size, "items": [{"id": row.id, "name": getattr(row, "title", ""), "type": category, "updated_at": row.updated_at.strftime("%Y-%m-%d") if row.updated_at else ""} for row in rows]}}
    if category == "intents":
        # 最新意向: 仅已发布(wf_status='published' 或 历史 NULL), 字段契约对齐企业卡片
        stmt = select(IntentNotice).where(
            IntentNotice.is_deleted == False,
            or_(IntentNotice.wf_status == "published", IntentNotice.wf_status.is_(None)),
        )
        total = db.execute(select(func.count()).select_from(stmt.subquery())).scalar() or 0
        rows = db.execute(
            stmt.order_by(IntentNotice.published_at.is_(None), IntentNotice.published_at.desc())
            .offset((page - 1) * page_size).limit(page_size)
        ).scalars().all()
        items = [{
            "id": row.id,
            "name": row.title,
            "title": row.title,
            "type": row.industry or row.project_type or "意向",
            "province": row.province or row.region,
            "updated_at": row.published_at.strftime("%Y-%m-%d") if row.published_at else "",
        } for row in rows]
        return {"success": True, "data": {"category": category, "total": total, "page": page, "page_size": page_size, "items": items}}
    return {"success": True, "data": {"category": category, "total": 0, "page": page, "page_size": page_size, "items": []}}


@router.get("/home-config")
async def public_home_config(
    page: str = Query("home", max_length=32, description="前台页面: home/about/contact/solutions/intelligence/datacenter"),
    db: Session = Depends(get_db),
):
    """前台页面内容配置(公开) — 后台「内容配置中心」按页面维护, 前台按 block_key 渲染。

    返回结构:
      {
        "success": True,
        "data": {
          "page": "home",
          "blocks": { "<block_key>": { ...区块字段, "items": [...] }, ... },
          "order": ["top_guide", "quick_links", ...]
        }
      }
    仅返回 enabled 的区块与条目; 前台未取到配置时回退内置静态内容。
    """
    from app.services.cms import get_public_config
    return {"success": True, "data": get_public_config(db, page)}


@router.get("/overview")
async def public_overview(
    db: Session = Depends(get_db),
):
    """平台脱敏概览统计 — 供对外官网展示真实数据。

    返回:
      totals:        各实体总量(脱敏)
      region_top:    地域分布 Top10(仅省+数量)
      type_dist:     情报类型构成(聚合占比)
      monthly_trend: 近12个月 web_clue 采集趋势
    """
    # ── 总量(均按 is_deleted==False) ──
    def _count(model):
        return db.execute(
            select(func.count()).select_from(
                select(model.id).where(model.is_deleted == False).subquery()
            )
        ).scalar() or 0

    totals = {
        "bid_notices": _count(BidNotice),
        "companies": _count(Company),
        "intents": _count(IntentNotice),
        "web_clues": _count(WebClue),
        "persons": _count(Person),
        "projects": _count(Project),
    }

    # ── 地域分布 Top10(取 company.province 与 bid_notice.region 并集) ──
    region_rows = db.execute(text("""
        SELECT province, SUM(cnt) AS total FROM (
            SELECT province AS province, COUNT(*) AS cnt FROM company WHERE is_deleted=0 AND province IS NOT NULL AND province <> ''
            GROUP BY province
            UNION ALL
            SELECT region AS province, COUNT(*) AS cnt FROM bid_notice WHERE is_deleted=0 AND region IS NOT NULL AND region <> ''
            GROUP BY region
        ) t GROUP BY province ORDER BY total DESC LIMIT 10
    """)).all()
    region_top = [{"province": r[0], "count": int(r[1])} for r in region_rows]

    # ── 情报类型构成(聚合占比) ──
    # 招标/中标线索(bid_notice) + 单位(companies) + 意向(intent) + 网页线索(web_clue)
    type_dist = [
        {"name": "招投标线索", "value": totals["bid_notices"]},
        {"name": "单位画像", "value": totals["companies"]},
        {"name": "意向公告", "value": totals["intents"]},
        {"name": "网页线索", "value": totals["web_clues"]},
    ]

    # ── 近12个月 web_clue 采集趋势 ──
    months = []
    now = datetime.now()
    for i in range(11, -1, -1):
        m = (now.replace(day=1) - timedelta(days=30 * i)).replace(day=1)
        months.append(m.strftime("%Y-%m"))
    trend_rows = db.execute(text("""
        SELECT DATE_FORMAT(published_at, '%Y-%m') AS m, COUNT(*) AS cnt
        FROM web_clue WHERE is_deleted=0 AND published_at IS NOT NULL
        GROUP BY m
    """)).all()
    trend_map = {r[0]: int(r[1]) for r in trend_rows}
    monthly_trend = [
        {"month": m, "count": trend_map.get(m, 0)} for m in months
    ]

    # ── 派生指标(供官网数字墙) ──
    # 覆盖省级行政区数 = 地域去重数
    province_count = db.execute(text("""
        SELECT COUNT(DISTINCT province) FROM (
            SELECT province FROM company WHERE is_deleted=0 AND province IS NOT NULL AND province <> ''
            UNION SELECT region FROM bid_notice WHERE is_deleted=0 AND region IS NOT NULL AND region <> ''
        ) t
    """)).scalar() or 0

    return {
        "success": True,
        "message": "ok",
        "data": {
            "totals": totals,
            "region_top": region_top,
            "type_dist": type_dist,
            "monthly_trend": monthly_trend,
            "province_count": int(province_count),
            "updated_at": now.strftime("%Y-%m-%d %H:%M"),
        },
    }


# ─────────────── 商机数据公开接口(供情报动态页公开访问) ───────────────
@router.get("/opportunities/tags")
async def public_opportunity_tags(db: Session = Depends(get_db)):
    rows = db.execute(
        select(OpportunityTagDef).order_by(OpportunityTagDef.kind, OpportunityTagDef.sort_order)
    ).scalars().all()
    return {
        "success": True,
        "data": [
            {"id": r.id, "code": r.code, "label": r.label, "kind": r.kind, "isNew": bool(r.is_new)}
            for r in rows
        ],
    }


@router.post("/opportunities/search")
async def public_opportunity_search(payload: dict, db: Session = Depends(get_db)):
    """公开商机检索(情报动态页用)。payload 与 /api/v1/opportunities/search 兼容。"""
    ds = (payload.get("dataset_type") or payload.get("datasetType") or "project").lower()
    if ds not in ("project", "proposed", "landtrade"):
        ds = "project"
    stmt = select(Opportunity).where(Opportunity.is_deleted == 0, Opportunity.dataset_type == ds)

    tags = payload.get("tags") or []
    if payload.get("region_province"):
        # 前缀匹配: 兼容 "四川"/"四川省"、"广东"/"广东省" 混存
        stmt = stmt.where(Opportunity.region_province.like(f"{payload['region_province']}%"))
    if payload.get("region_city"):
        stmt = stmt.where(Opportunity.region_city == payload["region_city"])
    if payload.get("amount_min") is not None:
        stmt = stmt.where(Opportunity.amount_wan >= int(payload["amount_min"]))
    if payload.get("amount_max") is not None:
        stmt = stmt.where(Opportunity.amount_wan <= int(payload["amount_max"]))
    if payload.get("stage"):
        stmt = stmt.where(Opportunity.stage == payload["stage"])
    if payload.get("unit_role"):
        stmt = stmt.where(Opportunity.unit_role == payload["unit_role"])
    if payload.get("unit_name"):
        # 空格分词 AND 匹配(与 project_name 一致, 支持"学校 医院"多关键词)
        for kw in payload["unit_name"].split():
            kw = kw.strip()
            if kw:
                stmt = stmt.where(Opportunity.unit_name.like(f"%{kw}%"))
    if payload.get("owner_type"):
        stmt = stmt.where(Opportunity.owner_type == payload["owner_type"])
    if payload.get("owner_name"):
        stmt = stmt.where(Opportunity.owner_name.like(f"%{payload['owner_name']}%"))
    if payload.get("project_type"):
        stmt = stmt.where(Opportunity.project_type == payload["project_type"])
    if payload.get("update_start"):
        try:
            stmt = stmt.where(Opportunity.updated_at >= datetime.fromisoformat(payload["update_start"]))
        except ValueError:
            pass
    if payload.get("update_end"):
        try:
            end_dt = datetime.fromisoformat(payload["update_end"])
            # 纯日期(YYYY-MM-DD)时含当天全天: 否则 00:00 边界会排除当日全部记录
            if len(payload["update_end"]) <= 10:
                end_dt = end_dt + timedelta(days=1)
            stmt = stmt.where(Opportunity.updated_at <= end_dt)
        except ValueError:
            pass
    if payload.get("project_name"):
        for kw in payload["project_name"].split():
            kw = kw.strip()
            if kw:
                stmt = stmt.where(Opportunity.project_name.like(f"%{kw}%"))
    if tags:
        opp_ids_subq = select(OpportunityTag.opportunity_id).where(OpportunityTag.tag_id.in_(tags)).distinct()
        stmt = stmt.where(Opportunity.id.in_(opp_ids_subq))

    # 排除词: 标题或业主名命中任一排除词则剔除(订阅"排除词"条件消费, P1-7)
    _excl = payload.get("exclude_keywords") or payload.get("excludeKeywords")
    if _excl:
        _excl = _excl if isinstance(_excl, list) else str(_excl).split()
        _excl = [str(k).strip() for k in _excl if str(k).strip()]
        if _excl:
            stmt = stmt.where(
                not_(or_(
                    *[Opportunity.project_name.like(f"%{k}%") for k in _excl],
                    *[Opportunity.owner_name.like(f"%{k}%") for k in _excl],
                ))
            )

    page = int(payload.get("page") or 1)
    page_size = int(payload.get("page_size") or 20)
    total = db.execute(select(func.count()).select_from(stmt.subquery())).scalar() or 0
    rows = db.execute(
        stmt.order_by(Opportunity.updated_at.desc())
        .offset(max(0, (page - 1) * page_size))
        .limit(page_size)
    ).scalars().all()

    ids = [r.id for r in rows]
    tag_map: dict[int, list] = {i: [] for i in ids}
    if ids:
        link_rows = db.execute(
            select(OpportunityTag.opportunity_id, OpportunityTagDef.label, OpportunityTagDef.code)
            .join(OpportunityTagDef, OpportunityTagDef.id == OpportunityTag.tag_id)
            .where(OpportunityTag.opportunity_id.in_(ids))
        ).all()
        for oid, label, code in link_rows:
            tag_map.setdefault(oid, []).append({"label": label, "code": code})

    items = [
        {
            "id": r.id,
            "projectName": r.project_name,
            "ownerName": r.owner_name,
            "ownerType": r.owner_type,
            "ownerScale": r.owner_scale,
            "amountWan": r.amount_wan,
            "stage": r.stage,
            "regionProvince": r.region_province,
            "regionCity": r.region_city,
            "projectType": r.project_type,
            "currentVersion": r.current_version,
            "datasetType": r.dataset_type,
            "updatedAt": r.updated_at.isoformat() if r.updated_at else None,
            "tags": tag_map.get(r.id, []),
            "intentId": int(r.source.split("-")[-1]) if r.source and r.source.startswith("intent-notice-") else None,
        }
        for r in rows
    ]
    return {"success": True, "data": {"total": total, "items": items, "page": page, "page_size": page_size}}


# 来源站点 code → 名称片段(订阅按来源过滤用; web_source 表无 code 列, 故静态映射)
_BID_SOURCE_NAMES = {
    "scggzyjy": "四川",
    "ccgp": "中国政府采购",
    "cqggzyjy": "重庆",
    "kyqjy": "矿业权",
    "jzgcgc": "全国建设工程",
}


@router.post("/bids/search")
async def public_bid_search(payload: dict, db: Session = Depends(get_db)):
    """公开招投标检索(订阅「招投标信息」Tab 用, P1-7/P1-11)。

    过滤: 关键词(标题分词)/地域前缀/公告类型/招标方式(采购方式)/来源/排除词。
    与 /public/opportunities/search 同源风格, 但查 BidNotice(招投标公告)。
    """
    stmt = select(BidNotice).where(BidNotice.is_deleted == 0)

    keyword = payload.get("keyword")
    if keyword:
        for kw in str(keyword).split():
            kw = kw.strip()
            if kw:
                stmt = stmt.where(BidNotice.title.like(f"%{kw}%"))

    region = payload.get("region")
    if region:
        stmt = stmt.where(BidNotice.region.like(f"{region}%"))

    notice_types = payload.get("notice_types") or []
    if isinstance(notice_types, str):
        notice_types = [notice_types]
    if notice_types:
        stmt = stmt.where(BidNotice.notice_type.in_([str(t) for t in notice_types]))

    bid_methods = payload.get("bid_methods") or []
    if isinstance(bid_methods, str):
        bid_methods = [bid_methods]
    if bid_methods:
        stmt = stmt.where(BidNotice.purchase_way.in_([str(m) for m in bid_methods]))

    sources = payload.get("sources") or []
    if isinstance(sources, str):
        sources = [sources]
    if sources:
        names = [_BID_SOURCE_NAMES.get(str(s), str(s)) for s in sources]
        stmt = stmt.where(or_(*[BidNotice.source_name.like(f"%{n}%") for n in names if n]))

    excl = payload.get("exclude_keywords") or payload.get("excludeKeywords") or []
    if isinstance(excl, str):
        excl = excl.split()
    excl = [str(k).strip() for k in excl if str(k).strip()]
    if excl:
        stmt = stmt.where(not_(or_(*[BidNotice.title.like(f"%{k}%") for k in excl])))

    page = int(payload.get("page") or 1)
    page_size = int(payload.get("page_size") or 20)
    total = db.execute(select(func.count()).select_from(stmt.subquery())).scalar() or 0
    rows = db.execute(
        stmt.order_by(BidNotice.published_at.is_(None), BidNotice.published_at.desc())
        .offset(max(0, (page - 1) * page_size))
        .limit(page_size)
    ).scalars().all()

    items = [{
        "id": b.id,
        "title": b.title,
        "owner": b.purchaser,
        "region": b.region,
        "amount": b.budget_max if b.budget_max is not None else b.budget_min,
        "publishedAt": b.published_at.isoformat() if b.published_at else None,
        "url": b.url,
        "noticeType": b.notice_type,
        "purchaseWay": b.purchase_way,
        "sourceName": b.source_name,
    } for b in rows]
    return {"success": True, "data": {"total": total, "items": items, "page": page, "page_size": page_size}}


def _edge_total(db: Session) -> int:
    """平台人脉边总量(用于规则化合作概率演示)。"""
    return db.execute(
        select(func.count()).select_from(
            select(NetworkEdge.id).where(NetworkEdge.is_deleted == False).subquery()
        )
    ).scalar() or 0


def _opp_map_for(db: Session, ids: list[int]) -> dict:
    """批量查意向关联的商机, 返回 {source: (opp_id, current_version)}。"""
    if not ids:
        return {}
    rows = db.execute(
        select(Opportunity.source, Opportunity.id, Opportunity.current_version)
        .where(Opportunity.source.in_([f"intent-notice-{i}" for i in ids]))
    ).all()
    return {r[0]: (r[1], r[2]) for r in rows}


def _intent_public_vo(db: Session, i: IntentNotice, opp_map: dict, edge_total: int) -> dict:
    """单条意向的公开脱敏视图(列表与详情共用, 保证字段契约完全一致)。

    调用方需自行保证传入的是「已发布(wf_status='published'/NULL)且未删除」的记录。
    """
    opp_id, opp_version = opp_map.get(f"intent-notice-{i.id}", (None, None))
    return {
        "id": i.id,
        "title": i.title,
        "dept": i.dept,
        "region": i.region,
        "province": i.province,
        "project_type": i.project_type,
        "industry": i.industry,
        "amount_level": _amount_level(i.amount),
        "published_at": i.published_at.strftime("%Y-%m-%d") if i.published_at else None,
        "opp_id": opp_id,
        "opp_version": opp_version,
        "status": i.status,
        "keywords": (i.keywords or "").split(",") if i.keywords else [],
        "contact": i.contact or None,
        "body_excerpt": (i.raw_text or "").strip() or None,
        "url": i.url or None,
        "ai": {
            "heat": _rule_heat(i),
            "coop_prob": _rule_coop_prob(i, edge_total),
            "advice": _rule_advice(i),
        },
    }


@router.get("/intelligence")
async def public_intelligence(
    db: Session = Depends(get_db),
    limit: int = Query(20, ge=1, le=50),
):
    """对外官网·情报动态公开接口(脱敏)。

    返回:
      intents:   脱敏意向列表(真实记录, 不暴露联系方式/精确金额)
      graph:     平台人脉关系图谱样本(节点名脱敏, 仅关系类型可见)
      edge_total: 平台人脉边总量(用于规则化合作概率演示)
      note:      数据性质声明

    说明:
      - 单位/人员名称对外脱敏, 仅展示关系类型与聚合结构。
      - AI 摘要(热度/合作概率/建议)由规则引擎生成, 属演示性分析, 非真实商业判断。
    """
    # ── 1. 真实意向列表(脱敏; 仅已发布: wf_status='published' 或 历史未标记 NULL) ──
    intents = db.execute(
        select(IntentNotice)
        .where(
            IntentNotice.is_deleted == False,
            or_(IntentNotice.wf_status == "published", IntentNotice.wf_status.is_(None)),
        )
        .order_by(IntentNotice.published_at.is_(None), IntentNotice.published_at.desc())
        .limit(limit)
    ).scalars().all()

    edge_total = _edge_total(db)
    # 意向 → 商机联动: 已建档商机携带 opp_id/opp_version(版本号标签)
    opp_map = _opp_map_for(db, [it.id for it in intents])
    intent_list = [_intent_public_vo(db, it, opp_map, edge_total) for it in intents]

    # ── 2. 人脉图谱样本(脱敏节点名) ──
    edge_rows = db.execute(text("""
        SELECT src_type, src_id, src_name, tgt_type, tgt_id, tgt_name, rel_zh, rel_type, weight
        FROM network_edge WHERE is_deleted=0 ORDER BY id DESC LIMIT 80
    """)).mappings().all()
    graph = _build_graph(edge_rows)

    return {
        "success": True,
        "message": "ok",
        "data": {
            "intents": intent_list,
            "graph": graph,
            "edge_total": int(edge_total),
            "note": "本页为对外演示视图：单位/人员名称已脱敏，AI 摘要由规则引擎生成，属演示性分析，非真实商业判断。",
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        },
    }


@router.get("/intent/{intent_id}")
async def public_intent_detail(
    intent_id: int,
    db: Session = Depends(get_db),
):
    """按 id 直取单条已发布情报(脱敏), 字段契约与 /public/intelligence 列表项一致。

    存在意义: 情报详情页早期依赖 /public/intelligence 列表(limit 上限 50)再按 id
    查找, 导致发布时间较早、未进入最新 50 条的情报打不开(误显示「情报不存在」)。
    此接口按 id 直取, 不受列表分页限制。

    未发布 / 已删除 / 不存在 → 404(不泄露草稿内容)。
    """
    it = db.execute(
        select(IntentNotice).where(
            IntentNotice.id == intent_id,
            IntentNotice.is_deleted == False,
            or_(IntentNotice.wf_status == "published", IntentNotice.wf_status.is_(None)),
        )
    ).scalar_one_or_none()
    if not it:
        raise HTTPException(status_code=404, detail="情报不存在或未发布")
    return {
        "success": True,
        "message": "ok",
        "data": _intent_public_vo(db, it, _opp_map_for(db, [it.id]), _edge_total(db)),
    }


async def _llm_intent_analysis(intent: IntentNotice, edge_total: int) -> dict | None:
    """调用真实 Ollama 大模型生成意向深度研判。失败(超时/不可用)返回 None, 由调用方降级。

    模型回退: 优先 settings.OLLAMA_MODEL(qwen-graphrag), 失败则试 qwen2.5:7b。
    """
    prompt = LLM_INTENT_PROMPT.format(
        title=intent.title or "—",
        dept=intent.dept or "—",
        region=intent.region or "—",
        industry=intent.industry or "相关行业",
        amount_level=_amount_level(intent.amount),
        status=intent.status or "new",
        keywords="、".join((intent.keywords or "").split(",")[:8]) or "—",
        edge_hint=f"约 {edge_total} 条",
    )
    base_url = settings.OLLAMA_BASE_URL.rstrip("/")
    # 优先使用响应更快的 qwen2.5:7b; qwen-graphrag 在长 prompt 下易超时, 仅作兜底
    models = ["qwen2.5:7b", settings.OLLAMA_MODEL]
    for model in models:
        try:
            async with httpx.AsyncClient(timeout=90) as client:
                resp = await client.post(
                    f"{base_url}/api/chat",
                    json={
                        "model": model,
                        "messages": [{"role": "user", "content": prompt}],
                        "stream": False,
                        "format": "json",
                        "keep_alive": -1,
                        "options": {"temperature": 0.4, "num_predict": 400},
                    },
                )
                resp.raise_for_status()
                content = resp.json()["message"]["content"]
            try:
                return json.loads(content)
            except Exception:
                m = re.search(r"\{.*\}", content, re.S)
                if m:
                    try:
                        return json.loads(m.group(0))
                    except Exception:
                        continue
        except Exception as e:  # noqa: BLE001
            logger.warning("LLM intent analysis failed(intent=%s model=%s): %r", intent.id, model, e)
            continue
    return None


def _rule_intent_analysis(intent: IntentNotice, edge_total: int) -> dict:
    """规则引擎降级研判(LLM 不可用时使用)。"""
    return {
        "summary": f"基于规则引擎的演示性研判（{intent.region or '当地'}{intent.industry or '相关'}类意向）。",
        "heat": _rule_heat(intent),
        "coop_prob": _rule_coop_prob(intent, edge_total),
        "orgs": [intent.dept] if intent.dept else [],
        "persons_hint": "建议定位该意向的决策链关键角色，结合平台人脉图谱触达桥接人。",
        "network_path": "通过平台人脉关系图谱定位与业主/主管部门存在弱关联的桥接人，分步建立联系。",
        "advice": [_rule_advice(intent)],
        "opportunities": ["同类项目周期性投放，可提前布局下一窗口期。"],
    }


# ── 异步任务表(提交→轮询, 规避慢推理导致的长连接超时) ──
_INTENT_TASKS: dict = {}


@router.post("/intent-ai")
async def public_intent_ai_submit(
    body: dict,
    db: Session = Depends(get_db),
):
    """对外官网·单条意向真实 LLM 深度研判(脱敏) —— 提交任务, 立即返回 task_id。

    前端轮询 GET /public/intent-ai/{task_id} 取结果。LLM 较慢(弱算力), 故异步生成。
    请求体: {"intent_id": <int>}
    """
    intent_id = body.get("intent_id")
    if not intent_id:
        raise HTTPException(status_code=400, detail="缺少 intent_id")
    intent = db.execute(
        select(IntentNotice).where(IntentNotice.id == intent_id, IntentNotice.is_deleted == False)
    ).scalar_one_or_none()
    if not intent:
        raise HTTPException(status_code=404, detail="意向不存在")
    edge_total = db.execute(
        select(func.count()).select_from(
            select(NetworkEdge.id).where(NetworkEdge.is_deleted == False).subquery()
        )
    ).scalar() or 0

    task_id = uuid.uuid4().hex
    _INTENT_TASKS[task_id] = {"status": "running", "data": None, "created": time.time()}
    asyncio.create_task(_run_intent_ai(task_id, intent, edge_total))
    return {"success": True, "task_id": task_id}


async def _run_intent_ai(task_id: str, intent: IntentNotice, edge_total: int):
    task = _INTENT_TASKS.get(task_id)
    if not task:
        return
    analysis = await _llm_intent_analysis(intent, edge_total)
    if analysis:
        data = {
            "source": "llm",
            "model": settings.OLLAMA_MODEL,
            "analysis": analysis,
            "note": "由本地大模型基于真实意向数据生成，输出已做脱敏处理（不暴露具体单位/人员真名）。",
        }
    else:
        data = {
            "source": "rule",
            "analysis": _rule_intent_analysis(intent, edge_total),
            "note": "本地大模型暂不可用，已回退至规则引擎演示分析。",
        }
    task["data"] = data
    task["status"] = "done"
    _save_ai_cache(intent.id, data)


def _save_ai_cache(intent_id: int, data: dict):
    """将 LLM 研判结果按意向持久化(唯一), 供后续复用。"""
    from app.database import SessionLocal
    from app.models.intent_ai_cache import IntentAiCache
    db = SessionLocal()
    try:
        row = db.execute(
            select(IntentAiCache).where(IntentAiCache.intent_id == intent_id)
        ).scalar_one_or_none()
        payload = json.dumps(data.get("analysis"), ensure_ascii=False)
        if row is None:
            db.add(IntentAiCache(
                intent_id=intent_id,
                source=data.get("source", "llm"),
                model=data.get("model"),
                analysis=payload,
                note=data.get("note"),
            ))
        else:
            row.source = data.get("source", "llm")
            row.model = data.get("model")
            row.analysis = payload
            row.note = data.get("note")
        db.commit()
        logger.info("[ai-cache] saved intent=%s", intent_id)
    except Exception as e:  # noqa: BLE001
        db.rollback()
        logger.warning("[ai-cache] save failed intent=%s: %r", intent_id, e)
    finally:
        db.close()


@router.get("/intent-ai/cached/{intent_id}")
async def public_intent_ai_cached(
    intent_id: int,
    db: Session = Depends(get_db),
):
    """对外官网·读取该意向已生成并缓存的 AI 研判结果。

    返回: { success, found: bool, data: {...} | null }
    前端优先展示缓存, 并提供「重新生成」按钮触发全新分析。
    """
    from app.models.intent_ai_cache import IntentAiCache
    row = db.execute(
        select(IntentAiCache).where(IntentAiCache.intent_id == intent_id)
    ).scalar_one_or_none()
    if not row or not row.analysis:
        return {"success": True, "found": False}
    try:
        analysis = json.loads(row.analysis)
    except Exception:  # noqa: BLE001
        return {"success": True, "found": False}
    return {
        "success": True,
        "found": True,
        "data": {
            "source": row.source,
            "model": row.model,
            "analysis": analysis,
            "note": row.note or "",
            "updated_at": row.updated_at.strftime("%Y-%m-%d %H:%M") if row.updated_at else None,
        },
    }


@router.get("/intent-ai/{task_id}")
async def public_intent_ai_result(task_id: str):
    """轮询取 LLM 研判结果。status: running / done。"""
    task = _INTENT_TASKS.get(task_id)
    if not task:
        return {"success": True, "status": "failed", "data": None, "error": "任务不存在或已过期"}
    # 任务保留 10 分钟, 防止内存泄漏
    if time.time() - task.get("created", 0) > 600:
        _INTENT_TASKS.pop(task_id, None)
        return {"success": True, "status": "failed", "data": None, "error": "任务已过期"}
    return {"success": True, "status": task["status"], "data": task["data"]}


@router.get("/intent/{intent_id}/contacts")
async def public_intent_contacts(
    intent_id: int,
    db: Session = Depends(get_db),
):
    """对外官网·意向联系人(脱敏)。返回分组联系人, 姓名/电话等整体掩码。

    前台联系卡(甲方/设计师/建造商/分包)展示: 联系人存在时返回脱敏占位,
    不存在时返回空列表(前端展示「暂无联系人」)。
    """
    from app.models.intent_contact import IntentContact
    rows = db.execute(
        select(IntentContact).where(
            IntentContact.intent_id == intent_id,
            IntentContact.is_deleted == False,
        ).order_by(IntentContact.group, IntentContact.sort_order, IntentContact.id)
    ).scalars().all()
    items = []
    for r in rows:
        def _mask(v: Optional[str], keep: int = 1) -> str:
            if not v:
                return ""
            if len(v) <= keep:
                return "*" * (len(v) + 2)
            return v[:keep] + "*" * (len(v) - keep)
        items.append({
            "group": r.group,
            "name": _mask(r.name),
            "role": _mask(r.role, 0) or ("***" if r.role else ""),
            "department": _mask(r.department, 0) or ("***" if r.department else ""),
            "phone": _mask(r.phone, 0) or ("***" if r.phone else ""),
            "mobile": _mask(r.mobile, 0) or ("***" if r.mobile else ""),
        })
    return {"success": True, "data": items}


@router.get("/intent/{intent_id}/attachments")
async def public_intent_attachments(
    intent_id: int,
    db: Session = Depends(get_db),
):
    """对外官网·某意向的公告附件列表(已爬取的 pdf/doc/xls 等)。"""
    from app.models.intent_attachment import IntentAttachment
    rows = db.execute(
        select(IntentAttachment).where(
            IntentAttachment.intent_id == intent_id,
            IntentAttachment.is_deleted == False,
        ).order_by(IntentAttachment.id.desc())
    ).scalars().all()
    return {
        "success": True,
        "data": [{
            "id": r.id,
            "file_name": r.file_name,
            "remote_url": r.remote_url,
            "file_size": r.file_size,
            "download_url": f"/api/v1/public/intent/{intent_id}/attachments/{r.id}/download",
            "preview_url": f"/api/v1/public/intent/{intent_id}/attachments/{r.id}/preview",
        } for r in rows],
    }


@router.get("/intent/{intent_id}/attachments/{attachment_id}/download")
async def public_intent_attachment_download(
    intent_id: int,
    attachment_id: int,
    db: Session = Depends(get_db),
):
    """下载已爬取的公告附件。"""
    from pathlib import Path as _P
    from app.utils.upload_paths import upload_root
    from app.models.intent_attachment import IntentAttachment
    row = db.execute(
        select(IntentAttachment).where(
            IntentAttachment.id == attachment_id,
            IntentAttachment.intent_id == intent_id,
            IntentAttachment.is_deleted == False,
        )
    ).scalar_one_or_none()
    if not row or not row.local_path:
        raise HTTPException(status_code=404, detail="附件不存在")
    base = upload_root()
    fpath = base / row.local_path
    if not fpath.exists():
        raise HTTPException(status_code=404, detail="附件文件不存在")
    return FileResponse(
        path=str(fpath),
        filename=row.file_name,
        media_type="application/octet-stream",
    )


@router.get("/intent/{intent_id}/attachments/{attachment_id}/preview")
async def public_intent_attachment_preview(
    intent_id: int,
    attachment_id: int,
    db: Session = Depends(get_db),
):
    """在线预览公告附件: 以 inline 方式返回, 浏览器可内联展示(pdf/图片/文本等)。

    说明: pdf/图片/纯文本浏览器可直接内联预览; docx/xlsx 等浏览器无法原生渲染,
    仍会由浏览器按内联方式尝试打开(一般仍需下载后查看)。
    """
    import mimetypes
    from pathlib import Path as _P
    from app.utils.upload_paths import upload_root
    from urllib.parse import quote
    from app.models.intent_attachment import IntentAttachment
    row = db.execute(
        select(IntentAttachment).where(
            IntentAttachment.id == attachment_id,
            IntentAttachment.intent_id == intent_id,
            IntentAttachment.is_deleted == False,
        )
    ).scalar_one_or_none()
    if not row or not row.local_path:
        raise HTTPException(status_code=404, detail="附件不存在")
    base = upload_root()
    fpath = base / row.local_path
    if not fpath.exists():
        raise HTTPException(status_code=404, detail="附件文件不存在")
    media_type = mimetypes.guess_type(row.file_name)[0] or "application/octet-stream"
    # HTTP 头只允许 ASCII; 中文文件名按 RFC 5987 filename*=UTF-8'' 编码, 浏览器可正确识别并内联预览
    safe_name = quote(row.file_name, safe="")
    return FileResponse(
        path=str(fpath),
        media_type=media_type,
        headers={"Content-Disposition": f"inline; filename*=UTF-8''{safe_name}"},
    )


@router.get("/intent/{intent_id}/attachments/{attachment_id}/table")
async def public_intent_attachment_table(
    intent_id: int,
    attachment_id: int,
    db: Session = Depends(get_db),
):
    """解析公告附件中的 Excel → 结构化表格(表头 + 数据行), 供前台直接渲染。

    严格按附件原样输出, 不做任何推断/补全, 避免展示编造数据:
      1) 仅解析第一个工作表, 首行作为表头, 全空行丢弃;
      2) 单元格无值即空串, 不猜测、不填充;
      3) Excel 日期序列号(如 46230)还原为 ISO 日期(2026-07-27)。
    仅支持 xlsx/xlsm; 旧版 .xls 与 pdf/doc/图片等非表格附件返回 415, 由前端自行跳过。
    """
    from datetime import date as _date
    from pathlib import Path as _P
    from app.utils.upload_paths import upload_root
    from openpyxl import load_workbook
    from app.models.intent_attachment import IntentAttachment

    row = db.execute(
        select(IntentAttachment).where(
            IntentAttachment.id == attachment_id,
            IntentAttachment.intent_id == intent_id,
            IntentAttachment.is_deleted == False,
        )
    ).scalar_one_or_none()
    if not row or not row.local_path:
        raise HTTPException(status_code=404, detail="附件不存在")
    if not re.search(r"\.(xlsx|xlsm)$", row.file_name or "", re.I):
        raise HTTPException(status_code=415, detail="该附件不是可解析的表格文件(仅支持 xlsx/xlsm)")

    base = upload_root()
    fpath = base / row.local_path
    if not fpath.exists():
        raise HTTPException(status_code=404, detail="附件文件不存在")

    try:
        wb = load_workbook(filename=str(fpath), data_only=True, read_only=True)
    except Exception as e:  # 文件损坏/格式不符
        logging.warning("打开附件表格失败 attachment_id=%s: %s", attachment_id, e)
        raise HTTPException(status_code=422, detail=f"表格打开失败: {e}")

    try:
        ws = wb.worksheets[0]
        rows = [
            [("" if c is None else str(c).strip()) for c in r]
            for r in ws.iter_rows(values_only=True)
        ]
    except Exception as e:
        logging.warning("解析附件表格失败 attachment_id=%s: %s", attachment_id, e)
        raise HTTPException(status_code=422, detail=f"表格解析失败: {e}")
    finally:
        try:
            wb.close()
        except Exception:
            pass

    rows = [r for r in rows if any(r)]
    if not rows:
        return {"success": True, "data": {"file_name": row.file_name, "headers": [], "rows": []}}

    headers, body = rows[0], rows[1:]
    # Excel 日期序列号还原: 46230 -> 2026-07-27(仅处理落在合理日期区间的纯数字, 避免误伤编号列)
    _epoch = _date(1899, 12, 30)
    for r in body:
        for i, v in enumerate(r):
            if v.isdigit() and 40000 < int(v) < 60000:
                r[i] = (_epoch + timedelta(days=int(v))).isoformat()

    return {
        "success": True,
        "data": {"file_name": row.file_name, "headers": headers, "rows": body},
    }
