"""Excel 导入导出 API"""
import datetime
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.project import Project
from app.models.person import Person
from app.models.company import Company
from app.models.field_meta import FieldMetadata
from app.middleware.auth import get_current_user, require_permission
from openpyxl import load_workbook
from app.services.excel_service import export_entity_to_excel, parse_import_excel
from app.services.cache_service import cache_service
from app.models.web_clue import WebClue
from types import SimpleNamespace

router = APIRouter(prefix="/excel", tags=["Excel导入导出"])

# 导入 Excel 列头别名 → 模型字段(支持动态字段 ext_attrs)
COLUMN_ALIASES = {
    "甲方单位名称": "name",
    "单位名称": "name",
    "甲方单位类型": "company_type",
    "单位类型": "company_type",
    "甲方纳税人代码": "credit_code",
    "纳税人代码": "credit_code",
    "统一社会信用代码": "credit_code",
    "行业类别": "industry",
    "行业": "industry",
    "合同金额": "ext:contract_amount",
    "甲方联系方式": "ext:contact",
    "联系方式": "ext:contact",
    "省份": "province",
    "城市": "city",
    "地址": "address",
    "官网": "website",
    "简称": "short_name",
}


@router.post("/export/{entity_type}")
async def export_entity(
    entity_type: str,
    db: Session = Depends(get_db),
    user: dict = Depends(require_permission("api_excel")),
):
    """
    导出实体数据为 Excel（列由元数据驱动）

    请求示例:
      POST /api/v1/excel/export/project

    返回: .xlsx 文件流
    """
    if entity_type not in ("projects", "persons", "project_members", "companies", "web_clues"):
        raise HTTPException(status_code=400, detail="不支持的导出实体类型")

    # 获取字段元数据
    meta_collection = await cache_service.get_field_meta_list(entity_type.rstrip("s"))
    if not meta_collection:
        meta_objs = db.execute(
            select(FieldMetadata).where(
                FieldMetadata.entity_type == entity_type.rstrip("s"),
                FieldMetadata.is_exportable == True,
                FieldMetadata.status == "enabled",
                FieldMetadata.is_deleted == False,
            )
        ).scalars().all()
    else:
        from app.services.dynamic_field_engine import FieldMetadata as FM
        meta_objs = [FM(**m) for m in meta_collection]

    # 获取实体数据(数据范围过滤: 未启用时保持现状不过滤)
    from app.services.data_scope_service import resolve_scope, scope_filter
    if entity_type == "projects":
        stmt = select(Project).where(Project.is_deleted == False)
        scope = resolve_scope(db, user, "project")
        cond = scope_filter(scope, Project, "project", dept_id_col=Project.department_id)
        if cond is not None:
            stmt = stmt.where(cond)
        items = db.execute(stmt).scalars().all()
    elif entity_type == "persons":
        stmt = select(Person).where(Person.is_deleted == False)
        scope = resolve_scope(db, user, "person")
        cond = scope_filter(scope, Person, "person", dept_id_col=Person.department_id)
        if cond is not None:
            stmt = stmt.where(cond)
        items = db.execute(stmt).scalars().all()
    elif entity_type == "companies":
        stmt = select(Company).where(Company.is_deleted == False)
        scope = resolve_scope(db, user, "company")
        cond = scope_filter(scope, Company, "company")
        if cond is not None:
            stmt = stmt.where(cond)
        items = db.execute(stmt).scalars().all()
    elif entity_type == "web_clues":
        items = db.execute(
            select(WebClue).where(WebClue.is_deleted == False).order_by(WebClue.published_at.desc())
        ).scalars().all()
        # web_clue 无 field_metadata 种子, 在此直接构造导出列(避免导出空表)
        meta_objs = [
            SimpleNamespace(field_key=k, display_name=n, is_exportable=True, status="enabled", sort_order=o, option_set_code=None, data_type=None)
            for k, n, o in [
                ("id", "ID", 0), ("title", "标题", 1), ("url", "链接", 2),
                ("source_name", "来源", 3), ("region", "地域", 4), ("category", "分类", 5),
                ("status", "状态", 6), ("published_at", "发布时间", 7), ("fetched_at", "抓取时间", 8),
            ]
        ]
    else:
        items = []

    # 转 dict
    item_dicts = []
    for item in items:
        d = {c.name: getattr(item, c.name) for c in item.__table__.columns}
        # 展开 ext_attrs 以便 Excel 列映射
        if d.get("ext_attrs"):
            d.update(d["ext_attrs"])
        item_dicts.append(d)

    buf = export_entity_to_excel(db, entity_type.rstrip("s"), item_dicts, meta_objs)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={entity_type}_{timestamp}.xlsx"},
    )


@router.post("/import/{entity_type}")
async def import_entity(
    entity_type: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(require_permission("api_excel")),
):
    """
    导入 Excel（字段映射由元数据驱动，动态字段自动可导入）

    请求: multipart/form-data, file=xxx.xlsx

    响应示例:
      ```json
      {
        "success": true,
        "imported": 15,
        "skipped": 3,
        "errors": ["第5行动态字段校验失败: contract_amount: 输入值无效"]
      }
      ```
    """
    if entity_type not in ("projects", "persons", "companies"):
        raise HTTPException(status_code=400, detail="不支持的导入实体类型")

    # 统一读取文件 + 安全校验(扩展名/魔数, 防伪造类型上传)
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="上传文件为空")
    from app.utils.upload_security import check_upload_file
    err = check_upload_file(file_bytes, file.filename or "")
    if err:
        raise HTTPException(status_code=400, detail=err)

    # 人员走专用花名册导入(按姓名复用, 兼容业务列头; 通用映射会把部门文本塞进
    # department_id 整数列导致 500)
    if entity_type == "persons":
        from app.services.real_person_import import import_real_person
        try:
            result = import_real_person(db, file_bytes)
        except Exception as e:  # noqa: BLE001
            db.rollback()
            raise HTTPException(status_code=400, detail=f"导入失败: {e}") from e
        if not result.get("success"):
            raise HTTPException(status_code=400, detail=result.get("message", "导入失败"))
        return result

    # 项目合同完整导入: 公司/项目/负责人/业主/联系人/关联 + Neo4j 图谱 + 单位信息补全。
    # 模板见 samples/项目合同信息.xlsx(项目名称/法人单位/项目负责人/甲方单位名称/合同金额等)。
    # 列头不匹配时提示使用专用模板, 不再退回通用导入(通用导入不建图谱)。
    if entity_type == "projects":
        from app.services.real_project_import import import_real_project
        try:
            result = import_real_project(db, file_bytes)
        except Exception as e:  # noqa: BLE001
            db.rollback()
            raise HTTPException(status_code=400, detail=f"导入失败: {e}") from e
        if not result.get("success"):
            raise HTTPException(status_code=400, detail=result.get("message", "导入失败"))
        return result

    # 获取字段元数据
    entity_singular = entity_type.rstrip("s")
    meta_objs = db.execute(
        select(FieldMetadata).where(
            FieldMetadata.entity_type == entity_singular,
            FieldMetadata.is_deleted == False,
        )
    ).scalars().all()

    # 解析 Excel — 复用外层统一读取的文件内容(避免二次 read 读到空)
    xlsx_buf = io.BytesIO(file_bytes)

    imported = 0
    failed = 0
    skipped = 0
    all_errors = []
    rows = []
    parse_errors = []

    if entity_type == "companies":
        # company 样本为业务中文列头, 直接用别名映射解析, 不受元数据缺失影响
        try:
            wb = load_workbook(xlsx_buf, data_only=True)
        except Exception:  # 魔数通过但内容伪造/损坏
            raise HTTPException(status_code=400, detail="文件解析失败: 不是有效的 Excel 文件")
        ws = wb.active
        header_map = {}  # col_idx -> field_key / ext:field
        for col in range(1, ws.max_column + 1):
            hdr = (ws.cell(1, col).value or "").strip()
            alias = COLUMN_ALIASES.get(hdr)
            if alias:
                header_map[col] = alias
        for row_idx in range(2, ws.max_row + 1):
            builtin = {}
            ext = {}
            for col, alias in header_map.items():
                val = ws.cell(row_idx, col).value
                if val is None or str(val).strip() == "":
                    continue
                if alias.startswith("ext:"):
                    ext[alias[4:]] = val
                else:
                    builtin[alias] = val
            if not builtin and not ext:
                continue
            builtin["ext_attrs"] = ext
            rows.append(builtin)
    else:
        try:
            rows, parse_errors = parse_import_excel(xlsx_buf, meta_objs, db=db)
        except Exception:  # 魔数通过但内容伪造/损坏
            raise HTTPException(status_code=400, detail="文件解析失败: 不是有效的 Excel 文件")

    for e in parse_errors:
        all_errors.append({"row": 0, "field": "", "message": e})

    if entity_type == "projects":
        entity_cls = Project
    elif entity_type == "persons":
        entity_cls = Person
    elif entity_type == "companies":
        entity_cls = Company
    else:
        entity_cls = None

    if entity_cls is None:
        return {"success": False, "message": "unsupported entity type", "data": {"imported": 0, "failed": len(rows), "errors": all_errors}}

    imported_company_ids: list = []

    for idx, row in enumerate(rows, start=2):
        try:
            builtin_fields = {k: v for k, v in row.items() if k != "ext_attrs"}
            # Company.code 必填且唯一, 样本无此列时自动生成
            if entity_cls is Company and not builtin_fields.get("code"):
                raw_code = builtin_fields.get("credit_code")
                if raw_code and str(raw_code).strip() not in ("", "/", "-"):
                    builtin_fields["code"] = f"CO-{raw_code}"
                else:
                    ts = datetime.datetime.now().strftime("%y%m%d%H%M%S")
                    builtin_fields["code"] = f"AUTO-{ts}-{idx}"
            # Company.code 唯一, 重复公司(同一 credit_code 多行)跳过, 不计入失败
            if entity_cls is Company and builtin_fields.get("code"):
                exists = db.execute(
                    select(Company).where(Company.code == builtin_fields["code"], Company.is_deleted == False)
                ).scalars().first()
                if exists:
                    skipped += 1
                    continue
            instance = entity_cls(**builtin_fields)
            instance.ext_attrs = row.get("ext_attrs")
            db.add(instance)
            db.flush()
            db.commit()  # 逐行提交,避免单条失败回滚整批
            imported += 1
            if entity_cls is Company:
                imported_company_ids.append(instance.id)
        except Exception as e:
            db.rollback()  # 仅回滚当前行,已提交行不受影响
            failed += 1
            all_errors.append({"row": idx, "field": "", "message": str(e)})

    # 导入的公司同步 Neo4j 图谱节点(失败不阻断导入, 日志记录)
    for cid in imported_company_ids:
        try:
            from app.services.neo4j_sync import sync_company
            c = db.get(Company, cid)
            if c:
                sync_company(cid, c.name or "", c.code or "", c.company_type or "",
                             c.province or "", c.city or "")
        except Exception:  # noqa: BLE001
            pass

    return {
        "success": True,
        "message": "ok",
        "data": {
            "imported": imported,
            "failed": failed,
            "skipped": skipped,
            "errors": all_errors,
        },
    }


@router.get("/import/task/{task_id}")
async def get_import_task_status(task_id: str, db: Session = Depends(get_db),
                                 user: dict = Depends(get_current_user)):
    """查询导入后台任务状态(进度/日志/结果)。前端轮询此接口。"""
    from app.services.import_task import get_import_task
    task = get_import_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在或已过期")
    if task.get("user_id") != user["user_id"]:
        raise HTTPException(status_code=403, detail="无权查看该任务")
    return {"success": True, "data": task}
