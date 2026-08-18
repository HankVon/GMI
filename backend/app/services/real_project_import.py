"""真实项目 Excel 导入服务 — 完整导入(公司/人员/项目/关联/Neo4j 图谱)。

对标 samples/import_real_project.py 的完整导入流程, 用后端自身能力实现:
  1. 解析 xlsx (openpyxl)
  2. 法人单位公司(鑫冶) — 复用或创建
  3. 项目 — 复用或创建(状态=completed)
  4. 项目负责人(归属法人单位) — 创建并关联(role=项目负责人)
  5. 业主单位(甲方社区) — 复用或创建, 关联(role=owner)
  6. 业主联系人 — 创建并关联(role=业主联系人)
  7. 法人单位关联项目(constructor)
  8. 进度记录「项目已完工」 + Neo4j 同步

幂等: 已存在的公司/人员/项目/关联自动复用跳过, 可重复导入。
"""
import datetime
import io
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.company import Company, ProjectCompany
from app.models.person import Person
from app.models.project import Project
from app.models.project_member import ProjectMember
from app.models.project_progress import ProjectProgress
from app.services.neo4j_sync import (
    sync_project, sync_project_companies, sync_project_members, sync_company_colleagues,
)


_code_seq = 0


def _gen_code(prefix: str) -> str:
    """生成唯一业务编码: 秒级时间戳 + 自增序号, 短小可读且同秒/并发不撞唯一键。

    示例: EMP-IMP2608121107071 / EMP-IMP2608121107072
    """
    global _code_seq
    _code_seq += 1
    return f"{prefix}{datetime.datetime.now():%y%m%d%H%M%S}{_code_seq}"


def _parse_amount(s) -> int:
    """解析金额: 支持整数/小数/千分位逗号/空白。返回整数元(四舍五入)。"""
    if s is None:
        return 0
    s = str(s).strip().replace(",", "").replace("，", "").replace("元", "").strip()
    if not s or s.lower() == "nan":
        return 0
    try:
        return int(round(float(s)))
    except ValueError:
        return 0


# 行业类别 -> project_category 映射(规则优先, 未命中或空白用 AI 分析)
CATEGORY_RULES: list[tuple[tuple, str]] = [
    (("生态修复", "环境治理", "矿山修复", "土壤修复", "水污染治理", "植被恢复"), "eco_restoration"),
    (("地质灾害", "地灾", "滑坡", "崩塌", "泥石流", "地面沉降", "地灾防治"), "geo_hazard"),
    (("地质勘查", "地质调查", "勘察", "勘查", "测绘", "工程地质", "岩土"), "geo_survey"),
    (("矿业权", "采矿", "探矿", "矿山越界", "资源储量", "出让"), "mining_rights"),
    (("政策", "规划", "咨询", "评估"), "policy"),
]


def _map_category(text: str) -> str:
    """按行业类别文本映射 project_category 枚举值; 未命中返回空串。"""
    t = (text or "").strip()
    if not t or t.lower() == "nan":
        return ""
    for kws, cat in CATEGORY_RULES:
        if any(kw in t for kw in kws):
            return cat
    return ""


def _ai_classify(project_name: str, industry: str, description: str) -> str:
    """用本地 Ollama(qwen-graphrag) 分析项目所属分类。失败返回空串。"""
    try:
        import httpx
        candidates = "、".join(
            f"{v}：{v}" for v in
            ["eco_restoration", "policy", "geo_hazard", "geo_survey", "mining_rights"]
        )
        prompt = (
            f"请判断以下项目最可能属于哪个分类，只返回一个枚举值(eco_restoration/policy/geo_hazard/geo_survey/mining_rights)，不要解释。\n"
            f"项目名称:{project_name}\n行业类别:{industry or '（空）'}\n项目简介:{(description or '')[:300]}\n"
            f"可选分类: {candidates}"
        )
        resp = httpx.post("http://localhost:11434/api/generate", json={
            "model": "qwen-graphrag",
            "prompt": prompt,
            "stream": False,
            "options": {"temperature": 0.1},
        }, timeout=60)
        text = (resp.json() or {}).get("response") or ""
        for v in ("eco_restoration", "policy", "geo_hazard", "geo_survey", "mining_rights"):
            if v in text:
                return v
    except Exception:  # noqa: BLE001
        pass
    return ""


def _cell(row, col) -> str:
    """单元格安全取值(去空白, NaN 处理)。"""
    v = row.get(col)
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def _parse_ws(ws) -> list[dict]:
    """读取工作表首行为列头, 返回 list[dict]。"""
    headers = []
    for col in range(1, ws.max_column + 1):
        h = ws.cell(1, col).value
        headers.append(str(h).strip() if h else f"__col{col}__")
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for i, h in enumerate(headers, start=1):
            row[h] = ws.cell(r, i).value
        if any(v is not None and str(v).strip() not in ("", "nan") for v in row.values()):
            rows.append(row)
    return rows


def _find_company(db: Session, name: str) -> Optional[Company]:
    return db.execute(
        select(Company).where(Company.name == name, Company.is_deleted == False).limit(1)
    ).scalar_one_or_none()


def _find_person(db: Session, name: str, company_id: Optional[int] = None) -> Optional[Person]:
    """按姓名查人员(带单位消歧)。

    公告中的联系人常为「张老师/李工」这类称呼, 不同项目/单位可能同名却非同一人。
    若不区分单位, 仅按姓名查重会把两个无关项目的「唐老师」误合并成一个 person,
    导致所属单位错乱(见: 青白江水务局 vs 广安石笋镇政府 两个唐老师被合并)。

    消歧规则: 同名时, 仅当 company_id 相同 或 库中记录无单位(company_id 为空)才复用;
    否则视为不同人 → 返回 None 由调用方新建独立 person。
    """
    q = select(Person).where(Person.name == name, Person.is_deleted == False)
    if company_id:
        # 优先精确匹配: 同名 + 同单位
        exact = db.execute(q.where(Person.company_id == company_id).limit(1)).scalar_one_or_none()
        if exact:
            return exact
        # 退化为「无单位记录」可复用(历史导入的人员可能缺单位)
        return db.execute(q.where(Person.company_id.is_(None)).limit(1)).scalar_one_or_none()
    return db.execute(q.limit(1)).scalar_one_or_none()


def _find_project(db: Session, name: str) -> Optional[Project]:
    return db.execute(
        select(Project).where(Project.name == name, Project.is_deleted == False).limit(1)
    ).scalar_one_or_none()


# 占位符: 单位表历史数据用 "/" "-" "无" 等表示「无信息」, 需视为空
_PLACEHOLDERS = {"", "/", "-", "—", "无", "暂无", "null", "none", "nan"}


def _is_blank(v) -> bool:
    """字段是否有真实值: 空字符串 / None / 占位符(/, -, 无等) 都视为空白。"""
    if v is None:
        return True
    s = str(v).strip()
    return s.lower() in _PLACEHOLDERS


# 补全目标字段(对齐公司字段元数据): 任一空白即触发补全
_ENRICH_TARGET_FIELDS = (
    "legal_rep", "econ_kind", "registered_capital", "belong_org",
    "business_scope", "contact_person", "contact_phone",
    "establish_date", "oper_status", "reg_no", "contact",
)


def _enrich_company_from_qcc(db: Session, company, log: list, errors: list) -> None:
    """补全单位信息: 免费渠道优先(公告库/搜索引擎/政府采购网), 企查查兜底。

    判断「信息不全」: 省份/城市/地址/信用代码 + 11 个 ext 目标字段任一为空白
    (含 "/" "-" "无" 等占位符)即触发补全; 全部有真实值才跳过。
    注意: 历史数据常用 "/" 表示无信息, bool("/")=True 会导致误判跳过, 故用 _is_blank。
    """
    if company is None:
        return
    ext = company.ext_attrs or {}
    blanks = []
    if _is_blank(company.province):
        blanks.append("province")
    if _is_blank(company.city):
        blanks.append("city")
    if _is_blank(company.credit_code):
        blanks.append("credit_code")
    if _is_blank(company.address):
        blanks.append("address")
    for k in _ENRICH_TARGET_FIELDS:
        if _is_blank(ext.get(k)):
            blanks.append(k)
    if not blanks:
        return
    # 1) 免费渠道: 公告库匹配 + 主动检索四川政府采购网
    try:
        from app.services.company_free_enrich import enrich_company_free
        result = enrich_company_free(db, company)
        if result.get("ok"):
            src = result.get("source", "free")
            if result.get("updated"):
                log.append(f"免费补全[{company.name}]({src}): {'、'.join(result['updated'])}")
            else:
                log.append(f"免费补全[{company.name}]({src}): {result.get('message', '无可补全字段')}")
        else:
            log.append(f"免费补全[{company.name}]: {result.get('message', '查询失败')}")
    except Exception as e:  # noqa: BLE001
        errors.append(f"免费补全[{company.name}]失败: {e}")
    # 2) 企查查兜底(有配额时)
    try:
        from app.services.company_enrich import enrich_company_sync
        result = enrich_company_sync(company)
        if not result.get("ok"):
            if "未配置" not in (result.get("message") or ""):
                errors.append(f"企查查补全[{company.name}]: {result.get('message', '查询失败')}")
        elif result.get("updated"):
            log.append(f"企查查补全[{company.name}]: {'、'.join(result['updated'])}")
    except Exception as e:  # noqa: BLE001
        errors.append(f"企查查补全[{company.name}]失败: {e}")


def _add_project_company(db: Session, project_id: int, company_id: int, role: str, joined_at: Optional[str]) -> None:
    """项目-单位关联(幂等: 已存在同项目+同单位则跳过)。"""
    exists = db.execute(
        select(ProjectCompany).where(
            ProjectCompany.project_id == project_id,
            ProjectCompany.company_id == company_id,
            ProjectCompany.role == role,
            ProjectCompany.is_deleted == False,
        ).limit(1)
    ).scalar_one_or_none()
    if exists:
        return
    db.add(ProjectCompany(
        project_id=project_id, company_id=company_id, role=role,
        joined_at=_parse_dt(joined_at), is_active=True,
    ))


def _add_project_member(db: Session, project_id: int, person_id: int, role: str,
                        responsibility: str = "", joined_at: Optional[str] = None) -> None:
    """项目-人员关联(幂等: 已存在同项目+同人员则跳过)。"""
    exists = db.execute(
        select(ProjectMember).where(
            ProjectMember.project_id == project_id,
            ProjectMember.person_id == person_id,
            ProjectMember.role == role,
            ProjectMember.is_deleted == False,
        ).limit(1)
    ).scalar_one_or_none()
    if exists:
        return
    db.add(ProjectMember(
        project_id=project_id, person_id=person_id, role=role,
        responsibility=responsibility or None,
        joined_at=_parse_dt(joined_at), is_active=True,
    ))


def _parse_dt(s: Optional[str]) -> Optional[datetime.datetime]:
    """解析 'YYYY-MM-DD' 或 'YYYY-MM-DD HH:MM:SS' 为 datetime。"""
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
        try:
            return datetime.datetime.strptime(s[:19], fmt)
        except ValueError:
            continue
    return None


def _ensure_company(db: Session, name: str, company_type: str = "施工",
                    province: str = "", city: str = "", industry: str = "",
                    log: Optional[list] = None) -> int:
    """复用或创建公司, 返回 id。"""
    comp = _find_company(db, name)
    if not comp:
        comp = Company(
            code=_gen_code("CO-IMP"),
            name=name,
            short_name=name[:8],
            company_type=company_type,
            province=province or None,
            city=city or None,
            industry=industry or None,
        )
        db.add(comp)
        db.flush()
        if log is not None:
            log.append(f"创建公司[{name}] id={comp.id}")
    else:
        if log is not None:
            log.append(f"复用公司[{name}] id={comp.id}")
    return comp.id


def _import_one_project(db: Session, proj_rows: list[dict], log: list, errors: list) -> None:
    """导入单个项目(可能含多份分项合同行): 项目+进度+负责人+业主+法人单位关联+Neo4j同步。

    法人单位按该项目行的「法人单位」独立解析(支持不同项目不同法人单位)。
    """
    project_name = proj_rows[0]["项目名称"]
    # 法人单位: 取该项目第一行的法人单位(支持多法人单位 Excel)
    legal_name = _cell(proj_rows[0], "法人单位")
    if not legal_name:
        errors.append(f"项目[{project_name}]缺少法人单位, 跳过")
        return
    legal_id = _ensure_company(db, legal_name, company_type="施工",
                               province=_cell(proj_rows[0], "省份") or "",
                               city=_cell(proj_rows[0], "城市") or "",
                               industry=_cell(proj_rows[0], "行业类别"), log=log)
    # 法人单位信息补全(企查查, 只填空字段)
    _enrich_company_from_qcc(db, db.get(Company, legal_id), log, errors)
    # 最早开工日期作为项目启动日期
    start_dates = [_cell(c, "项目开工日期") for c in proj_rows if _cell(c, "项目开工日期")]
    start_dt = _parse_dt(min(start_dates)) if start_dates else None

    # ---- 1. 项目(复用或创建, 状态=completed) ----
    project = _find_project(db, project_name)
    total_amount = sum(_parse_amount(_cell(c, "合同金额")) for c in proj_rows)
    contracts_summary = "\n".join(
        f"- {_cell(c, '甲方单位名称') or _cell(c, '项目业主')}: {_cell(c, '合同金额')}元, "
        f"负责人{_cell(c, '项目负责人')}, 业主联系人{_cell(c, '业主联系人')}({_cell(c, '业主联系人电话')}), 开工{_cell(c, '项目开工日期')}"
        for c in proj_rows
    )
    contact_names = "/".join(sorted({_cell(c, "项目负责人") for c in proj_rows if _cell(c, "项目负责人")}))
    r0 = proj_rows[0]
    desc = (
        f"项目获取方式:{_cell(r0, '项目获取方式')}; 服务方式:{_cell(r0, '服务方式')}; "
        f"经营模式:{_cell(r0, '经营模式')}; 资金来源:{_cell(r0, '资金来源')}; "
        f"项目级别:{_cell(r0, '项目级别')}; 核算单元:{_cell(r0, '核算单元')}。\n分项合同:\n{contracts_summary}"
    )
    # 项目分类: 行业类别规则映射 -> 空白/未命中用 AI 分析兜底
    industry = _cell(r0, "行业类别")
    category = _map_category(industry)
    if not category:
        category = _ai_classify(project_name, industry, desc)
        if category:
            log.append(f"AI 分类[{project_name[:20]}...] -> {category} (行业类别空白)")
    if not category:
        category = "geo_survey"  # 兜底默认
    if not project:
        project = Project(
            code=_gen_code("PRJ-"),
            name=project_name,
            status="completed",
            start_date=start_dt,
            end_date=_parse_dt("2025-12-31"),
            description=desc,
            ext_attrs={"amount": str(total_amount), "category": category, "contact": contact_names},
        )
        db.add(project)
        db.flush()
        log.append(f"创建项目[{project_name}] id={project.id}")
    else:
        project.status = "completed"
        project.description = desc
        # 复用项目也更新金额与分类(纠正旧数据)
        attrs = project.ext_attrs or {}
        attrs = {**attrs, "amount": str(total_amount), "category": category}
        project.ext_attrs = attrs
        db.flush()
        log.append(f"复用项目[{project_name}] id={project.id}")
    project_id = project.id

    # ---- 2. 进度记录「项目已完工」(幂等) ----
    progress_exists = db.execute(
        select(ProjectProgress).where(
            ProjectProgress.project_id == project_id,
            ProjectProgress.title == "项目已完工",
            ProjectProgress.is_deleted == False,
        ).limit(1)
    ).scalar_one_or_none()
    if not progress_exists:
        db.add(ProjectProgress(
            project_id=project_id, title="项目已完工",
            content=f"{project_name} 已完成, 共 {len(proj_rows)} 份分项合同(合计 {total_amount} 元)均已完成交付。",
            progress_date=_parse_dt("2025-12-31"), sort_order=0,
        ))
        log.append("进度[项目已完工]")

    # ---- 3. 项目负责人(归属法人单位, 支持多人) ----
    leader_seen = set()
    for c in proj_rows:
        ld = _cell(c, "项目负责人")
        if not ld or ld in leader_seen:
            continue
        leader_seen.add(ld)
        person = _find_person(db, ld)
        if not person:
            person = Person(
                code=_gen_code("EMP-IMP"),
                name=ld, phone=_cell(c, "项目负责人联系电话"),
                company_id=legal_id, position="项目负责人", status="active",
            )
            db.add(person)
            db.flush()
            log.append(f"创建项目负责人[{ld}] id={person.id}")
        else:
            log.append(f"复用人员[{ld}] id={person.id}")
        _add_project_member(db, project_id, person.id, "manager",
                            responsibility=f"项目负责人(电话 {_cell(c, '项目负责人联系电话')})",
                            joined_at=_cell(c, "项目开工日期"))

    # ---- 4. 业主单位 + 业主联系人 ----
    owner_seen = set()
    for c in proj_rows:
        owner_name = _cell(c, "甲方单位名称") or _cell(c, "项目业主")
        if not owner_name or owner_name in owner_seen:
            continue
        owner_seen.add(owner_name)
        owner = _find_company(db, owner_name)
        owner_ext = {
            "reg_no": _cell(c, "甲方纳税人代码"),
            "contact": _cell(c, "甲方联系方式"),
        }
        owner_ext = {k: v for k, v in owner_ext.items() if v}
        if not owner:
            owner = Company(
                code=_gen_code("CO-OWN"),
                name=owner_name,
                short_name=(_cell(c, "项目业主") or owner_name)[:8],
                company_type=_cell(c, "甲方单位类型") or "业主",
                province=_cell(c, "省份") or "",
                city=_cell(c, "城市") or "",
                address=_cell(c, "甲方地址") or None,
                ext_attrs=owner_ext or None,
            )
            db.add(owner)
            db.flush()
            log.append(f"创建业主单位[{owner_name}] id={owner.id}")
        else:
            # 复用: 已有字段保留, 补充 xlsx 有而库里空的字段
            changed = []
            addr = _cell(c, "甲方地址")
            if addr and not owner.address:
                owner.address = addr
                changed.append("地址")
            if owner_ext:
                attrs = dict(owner.ext_attrs or {})
                for k, v in owner_ext.items():
                    if v and not attrs.get(k):
                        attrs[k] = v
                        changed.append(k)
                if changed:
                    owner.ext_attrs = attrs
            log.append(f"复用业主单位[{owner_name}] id={owner.id}"
                       + (f" 补充: {'、'.join(changed)}" if changed else ""))
        # 业主单位信息补全(企查查, 只填空字段: 法定代表人/电话/地址/信用代码等)
        _enrich_company_from_qcc(db, owner, log, errors)
        _add_project_company(db, project_id, owner.id, "owner", _cell(c, "项目开工日期"))

        cname = _cell(c, "业主联系人")
        if cname:
            cperson = _find_person(db, cname)
            if not cperson:
                cperson = Person(
                    code=_gen_code("EMP-OWN"),
                    name=cname, phone=_cell(c, "业主联系人电话"),
                    company_id=owner.id, position="业主联系人", status="active",
                )
                db.add(cperson)
                db.flush()
                log.append(f"创建业主联系人[{cname}] id={cperson.id}")
            _add_project_member(db, project_id, cperson.id, "业主联系人",
                                responsibility=f"甲方联系人(电话 {_cell(c, '业主联系人电话')})",
                                joined_at=_cell(c, "项目开工日期"))

    # ---- 5. 法人单位关联项目(constructor) ----
    _add_project_company(db, project_id, legal_id, "constructor", _cell(r0, "项目开工日期"))

    db.flush()

    # ---- 6. Neo4j 图谱同步 ----
    try:
        _p_ext = project.ext_attrs or {}
        sync_project(project_id, project.name, code=project.code or "", status="completed",
                     category=_p_ext.get("category", "") if isinstance(_p_ext, dict) else "",
                     province=_p_ext.get("province", "") if isinstance(_p_ext, dict) else "",
                     city=_p_ext.get("city", "") if isinstance(_p_ext, dict) else "",
                     county=_p_ext.get("county", "") if isinstance(_p_ext, dict) else "")
        pcs = db.execute(select(ProjectCompany).where(
            ProjectCompany.project_id == project_id, ProjectCompany.is_deleted == False)).scalars().all()
        companies = []
        for pc in pcs:
            c = db.get(Company, pc.company_id)
            companies.append({"company_id": pc.company_id,
                              "name": c.name if c else "", "role": pc.role or ""})
        sync_project_companies(project_id, companies)
        pms = db.execute(select(ProjectMember).where(
            ProjectMember.project_id == project_id, ProjectMember.is_deleted == False)).scalars().all()
        members = []
        for pm in pms:
            per = db.get(Person, pm.person_id)
            comp = db.get(Company, per.company_id) if per and per.company_id else None
            members.append({
                "person_id": pm.person_id,
                "name": per.name if per else "",
                "role": pm.role or "",
                "company_id": per.company_id if per else None,
                "company_name": comp.name if comp else "",
            })
        sync_project_members(project_id, members)
        # 重建同事关系: 对涉及的各公司(法人单位+业主单位), 为该单位全部人员建 COLLEAGUE 边
        company_ids = {pc.company_id for pc in pcs} | {per.company_id for per in
                        (db.get(Person, pm.person_id) for pm in pms) if per and per.company_id}
        for cid in company_ids:
            if not cid:
                continue
            persons_in_comp = db.execute(
                select(Person).where(Person.company_id == cid, Person.is_deleted == False)
            ).scalars().all()
            if len(persons_in_comp) >= 2:
                sync_company_colleagues(cid, [
                    {"person_id": p.id, "name": p.name} for p in persons_in_comp
                ])
        log.append(f"Neo4j 图谱同步[{project_name[:20]}...]")
    except Exception as e:  # noqa: BLE001
        errors.append(f"Neo4j 同步失败[{project_name[:20]}]: {e}")


def import_real_project(db: Session, file_bytes: bytes, entity_type: str = "projects") -> dict:
    """解析 xlsx 并按「项目名称」分组, 逐个完整导入(公司/人员/项目/关联/图谱)。

    同一项目名的多行视为分项合同, 合并为同一项目; 不同项目名分别创建项目。
    """
    if entity_type not in ("projects", "project"):
        raise ValueError("不支持的导入类型")
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = _parse_ws(ws)
    if not rows:
        return {"success": False, "message": "Excel 无数据", "imported": 0, "errors": ["无数据"]}

    # 校验必需列
    if not _cell(rows[0], "项目名称") or not _cell(rows[0], "法人单位"):
        return {"success": False, "message": "缺少「项目名称」或「法人单位」列",
                "imported": 0, "errors": ["缺少必需列"]}

    log: list[str] = []
    errors: list[str] = []

    # ---- 1. 按「项目名称」分组 ----
    groups: dict = {}
    for c in rows:
        pname = _cell(c, "项目名称")
        if not pname:
            continue
        groups.setdefault(pname, []).append(c)

    # ---- 2. 每组独立导入(法人单位按各自项目行解析) ----
    for pname, proj_rows in groups.items():
        try:
            _import_one_project(db, proj_rows, log, errors)
        except Exception as e:  # noqa: BLE001
            db.rollback()
            errors.append(f"项目[{pname}]导入失败: {e}")
            log.append(f"项目[{pname}]导入失败: {e}")

    db.commit()

    return {
        "success": True,
        "message": f"导入完成, 共 {len(groups)} 个项目",
        "project_count": len(groups),
        "project_id": None,
        "project_name": list(groups.keys())[0] if groups else "",
        "imported": len(log),
        "log": log,
        "errors": errors,
    }
