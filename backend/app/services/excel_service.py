"""
Excel 导入导出服务 — 元数据驱动字段映射

设计原则：
  导出: 按字段元数据 is_exportable 动态组装列
  导入: 按字段元数据解析列头 → 动态字段自动可导入
"""
import io
from typing import Optional, BinaryIO
from sqlalchemy import select
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.models.field_meta import FieldMetadata
from app.models.option_set import OptionSet, OptionItem
from app.services.dynamic_field_engine import validate_ext_attrs, validate_with_option_sets


def _load_option_maps(db: Session, metas: list) -> dict[str, dict]:
    """按 option_set_code 批量加载 value→label 映射(select 字段导出转中文)。"""
    maps: dict[str, dict] = {}
    codes = {m.option_set_code for m in metas if getattr(m, "option_set_code", None)}
    for code in codes:
        os_ = db.execute(
            select(OptionSet).where(OptionSet.code == code, OptionSet.is_deleted == False)
        ).scalar_one_or_none()
        if os_:
            items = db.execute(
                select(OptionItem).where(
                    OptionItem.option_set_id == os_.id, OptionItem.is_deleted == False
                )
            ).scalars().all()
            maps[code] = {str(i.value): i.label for i in items}
    return maps


def export_entity_to_excel(
    db: Session,
    entity_type: str,
    items: list[dict],
    meta_list: list[FieldMetadata],
) -> io.BytesIO:
    """
    按元数据驱动导出 Excel

    参数:
      db:           数据库会话
      entity_type:  实体类型
      items:        数据行列表,每行为 dict(含内置字段+ext_attrs展开)
      meta_list:    该实体的字段元数据列表

    返回: 包含 .xlsx 数据的 BytesIO
    """
    # 过滤可导出字段
    export_meta = [m for m in meta_list if m.is_exportable and m.status == "enabled"]
    export_meta.sort(key=lambda m: m.sort_order)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{entity_type}导出"

    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 表头
    headers = []
    for meta in export_meta:
        headers.append(meta.display_name)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    option_maps = _load_option_maps(db, export_meta)
    for row_idx, item in enumerate(items, 2):
        for col_idx, meta in enumerate(export_meta, 1):
            # 优先取展开值,其次 ext_attrs
            val = item.get(meta.field_key)
            if val is None and item.get("ext_attrs"):
                val = item["ext_attrs"].get(meta.field_key)
            # select / multi_select: 值 → 中文标签
            if val is not None and meta.option_set_code and meta.option_set_code in option_maps:
                om = option_maps[meta.option_set_code]
                if meta.data_type == "multi_select" and isinstance(val, list):
                    val = "、".join(om.get(str(x), str(x)) for x in val)
                else:
                    val = om.get(str(val), val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # 自适应列宽
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_import_excel(
    file: BinaryIO,
    meta_list: list[FieldMetadata],
    db: Optional[Session] = None,
) -> tuple[list[dict], list[str]]:
    """
    按元数据解析导入 Excel

    参数:
      file:      上传的 .xlsx 文件对象
      meta_list: 字段元数据列表
      db:        可选数据库会话; 提供时启用选项集值校验(select/multi_select)

    返回:
      (成功解析的行列表, 错误消息列表)

    映射逻辑:
      - 以 Excel 列头匹配字段 display_name 或 field_key
      - 动态字段自动映射进 ext_attrs
      - 内置字段(非 ext_attrs)保留在顶层
    """
    BUILTIN_FIELDS = {"code", "name", "description", "status", "manager_id",
                      "start_date", "end_date", "department_id",
                      "email", "phone", "position", "entry_date", "resign_date"}

    meta_map = {}
    for m in meta_list:
        key = (m.display_name or "").strip()
        meta_map[key] = m
        meta_map[m.field_key] = m  # 同时支持字段标识匹配

    wb = load_workbook(file, data_only=True)
    ws = wb.active

    if ws.max_row < 2:
        return [], ["Excel文件无数据行"]

    # 解析表头
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        headers.append((val or "").strip())

    # 解析数据行
    rows = []
    errors = []

    for row_idx in range(2, ws.max_row + 1):
        row_data = {"ext_attrs": {}}
        row_errors = []

        for col_idx, header in enumerate(headers, 1):
            if not header:
                continue

            cell_val = ws.cell(row_idx, col_idx).value
            if cell_val is None:
                continue

            meta = meta_map.get(header)
            if not meta:
                # 元数据未匹配 → 检查是否是内置字段列名
                if header in BUILTIN_FIELDS:
                    row_data[header] = cell_val
                # 都不是 → 跳过
                continue

            if meta.field_key in BUILTIN_FIELDS:
                row_data[meta.field_key] = cell_val
            else:
                row_data["ext_attrs"][meta.field_key] = cell_val

        # 校验动态字段(提供 db 时启用选项集值校验, 与 CRUD 路径一致)
        if row_data["ext_attrs"]:
            if db is not None:
                ok, cleaned, err_msg = validate_with_option_sets(
                    meta.entity_type,
                    row_data["ext_attrs"],
                    meta_list,
                    cache_svc=None,
                    db=db,
                )
            else:
                ok, cleaned, err_msg = validate_ext_attrs(
                    meta.entity_type,
                    row_data["ext_attrs"],
                    meta_list,
                )
            if not ok:
                row_errors.append(f"第{row_idx}行动态字段校验失败: {err_msg}")
                row_data["ext_attrs"] = {}
            else:
                row_data["ext_attrs"] = cleaned

        if row_errors:
            errors.extend(row_errors)
        else:
            rows.append(row_data)

    return rows, errors
