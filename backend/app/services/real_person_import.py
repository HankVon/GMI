"""真实人员 Excel 导入服务 — 按姓名复用更新/创建, 只取业务有效字段。

与 real_project_import 配套: 项目导入创建的人员(position=项目负责人/业主联系人)缺少
职位/电话/部门等基础信息, 本模块用人事花名册 xlsx 补齐/修正。

Excel 列头(24列, 只映射其中 5 个有业务意义的字段):
  - 姓名         -> Person.name (人员主键, 按姓名复用)
  - 主岗         -> Person.position (职位; 备选: 职务级别/职称)
  - 手机号码     -> Person.phone (电话; 备选: 办公电话)
  - 所属单位     -> Company.name 按名匹配 -> Person.company_id
  - 所属部门     -> ext_attrs.department (部门表不存在, 存文本动态字段)

幂等: 同名(未删除)人员复用并更新字段, 不重复创建; 可重复导入。
"""
import datetime
import io
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.company import Company
from app.models.person import Person
from app.services.neo4j_sync import sync_company_colleagues, sync_person

# 导入用人员编码前缀(与项目导入 EMP-IMP 区分, 避免秒级时间戳撞唯一键)
_PREFIX = "EMP-HR"


def _gen_code() -> str:
    """生成唯一人员编码: EMP-HR + 毫秒时间戳 + 微秒, 避免同秒撞 person.uk_code。"""
    return f"{_PREFIX}{datetime.datetime.now():%y%m%d%H%M%S%f}"  # noqa: DTZ005


def _cell(row, col) -> str:  # pyright: ignore[reportMissingParameterType]
    """单元格安全取值(去空白, NaN 处理)。"""
    v = row.get(col)
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def _parse_ws(ws) -> list[dict]:  # pyright: ignore[reportMissingParameterType]
    """读取工作表首行为列头, 返回 list[dict]。"""
    headers = []
    for col in range(1, ws.max_column + 1):
        h = ws.cell(1, col).value
        headers.append(str(h).strip() if h else f"__col{col}__")
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for i, h in enumerate(headers, start=1):
            row[h] = ws.cell(r, i).value
        if any(v is not None and str(v).strip() not in ("", "nan") for v in row.values()):
            rows.append(row)
    return rows


# 花名册关键列(用于自动定位列头行: 兼容首行为标题行「人员列表」的情况)
_PERSON_KEYS = {"姓名", "主岗", "手机号码", "所属单位", "所属部门", "单位名称", "职位", "电话", "办公电话"}


def _cell_grid(file_bytes: bytes) -> list[list]:
    """读取工作表全部单元格为二维网格(自动适配 xls/xlsx)。"""
    head = file_bytes[:8]
    if head[:4] == b"\xd0\xcf\x11\xe0":  # OLE2 = .xls
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        return [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]


def _load_rows(file_bytes: bytes, key_columns: set[str] | None = None) -> list[dict]:
    """兼容 .xls(xlrd) 与 .xlsx(openpyxl) 的解析, 返回 list[dict]。

    自动定位列头行: 有些导出文件首行是标题(如「人员列表」), 真实列头在第二行。
    取前 5 行中命中关键列名最多的一行作为列头, 数据从其下一行开始。
    """
    key_columns = key_columns or _PERSON_KEYS
    grid = _cell_grid(file_bytes)
    if not grid:
        return []

    best_idx, best_score = 0, -1
    for i in range(min(5, len(grid))):
        score = sum(1 for c in grid[i] if str(c).strip() in key_columns)
        if score > best_score:
            best_idx, best_score = i, score

    headers = [str(c).strip() or f"__col{j}__" for j, c in enumerate(grid[best_idx])]
    rows = []
    for r in range(best_idx + 1, len(grid)):
        row = {headers[j]: grid[r][j] for j in range(len(headers))}
        if any(str(v).strip() not in ("", "nan") for v in row.values()):
            rows.append(row)
    return rows


def _find_person(db: Session, name: str):
    return db.execute(
        select(Person).where(Person.name == name, Person.is_deleted == False).limit(1)
    ).scalar_one_or_none()


def _find_company(db: Session, name: str):
    """按名称精确匹配公司; 未命中时退化为包含匹配(去空白/括号差异)。"""
    comp = db.execute(
        select(Company).where(Company.name == name, Company.is_deleted == False).limit(1)
    ).scalar_one_or_none()
    if comp:
        return comp
    # 包含匹配: 名称含公司关键字(如「鑫冶」命中「四川省鑫冶岩土工程有限公司」)
    return db.execute(
        select(Company).where(
            Company.name.contains(name), Company.is_deleted == False
        ).order_by(Company.id.desc()).limit(1)
    ).scalar_one_or_none()


def _ensure_company(db: Session, name: str, log: Optional[list] = None):
    """按名称找单位, 未命中时自动创建(事业单位/企业按后缀判定)。

    人员花名册的「所属单位」多为机关/事业单位(地质大队/中心/局等),
    库存量单位缺失时应自动创建, 否则人员将失去单位归属。
    """
    comp = _find_company(db, name)
    if comp:
        return comp
    # 判定类型: 含「公司/厂/集团」→ 企业, 否则 → 事业单位
    ctype = "事业单位"
    if any(k in name for k in ("公司", "厂", "集团", "事务所", "有限")):
        ctype = "企业"
    comp = Company(
        code=_gen_company_code(),
        name=name,
        company_type=ctype,
        ext_attrs={"source": "花名册导入"},
    )
    db.add(comp)
    db.flush()
    if log is not None:
        log.append(f"自动创建单位[{name}] type={ctype} id={comp.id}")
    return comp


def _gen_company_code() -> str:
    """生成单位编码: CO-HR + 毫秒时间戳。"""
    return f"CO-HR{datetime.datetime.now():%y%m%d%H%M%S%f}"  # noqa: DTZ005


def import_real_person(db: Session, file_bytes: bytes, progress=None) -> dict:
    """解析人员花名册 xlsx, 按姓名复用更新/创建人员。

    progress: 可选回调(stage, imported, updated, skipped, failed, log), 用于后台任务实时进度。
    返回:
      {
        "success": bool,
        "message": str,
        "imported": int,   # 新建人数
        "updated": int,    # 复用更新人数
        "skipped": int,    # 无姓名/已软删跳过的行
        "log": [str],
        "errors": [str],
      }
    """
    rows = _load_rows(file_bytes, _PERSON_KEYS)
    if not rows:
        return {"success": False, "message": "Excel 无数据", "imported": 0,
                "updated": 0, "skipped": 0, "log": [], "errors": ["无数据"]}

    # 校验必需列(姓名/主岗 任一即可, 其余字段缺失只记日志)
    first = rows[0]
    has_name = bool(_cell(first, "姓名"))
    if not has_name:
        return {"success": False, "message": "缺少「姓名」列", "imported": 0,
                "updated": 0, "skipped": 0, "log": [], "errors": ["缺少必需列: 姓名"]}

    log: list[str] = []
    errors: list[str] = []
    imported = 0
    updated = 0
    skipped = 0

    for idx, row in enumerate(rows, start=2):
        name = _cell(row, "姓名")
        if not name:
            skipped += 1
            continue

        position = _cell(row, "主岗") or _cell(row, "职位") or _cell(row, "职务级别") or _cell(row, "职称")
        phone = _cell(row, "手机号码") or _cell(row, "电话") or _cell(row, "办公电话")
        unit_name = _cell(row, "所属单位") or _cell(row, "单位名称")
        department = _cell(row, "所属部门") or _cell(row, "部门")

        company_id = None
        if unit_name:
            comp = _find_company(db, unit_name)
            if not comp:
                comp = _ensure_company(db, unit_name, log=log)
            if comp:
                company_id = comp.id

        # 扩展字段(非核心, 存 ext_attrs): 性别/职称/职务级别/类别/类型/出生日期/办公电话/入职/身份证/副岗
        ext = {}
        if department:
            ext["department"] = department
        for xlsx_key, attr_key in (
            ("性别", "gender"), ("职称", "title"), ("职务级别", "position_level"),
            ("人员类别", "person_category"), ("人员类型", "person_type"),
            ("出生日期", "birth_date"), ("办公电话", "office_phone"),
            ("入职时间", "entry_date"), ("身份证号码", "id_card"), ("副岗", "secondary_position"),
        ):
            v = _cell(row, xlsx_key)
            if v:
                ext[attr_key] = v

        person = _find_person(db, name)
        try:
            if not person:
                person = Person(
                    code=_gen_code(),
                    name=name,
                    phone=phone or None,
                    company_id=company_id,
                    position=position or None,
                    status="active",
                    ext_attrs=ext or None,
                )
                db.add(person)
                db.flush()
                imported += 1
                log.append(f"[{idx}] 创建[{name}] 职位={position} 电话={phone} "  # pyright: ignore[reportImplicitStringConcatenation]
                           f"单位id={company_id} 部门={department}")
            else:
                changed = []
                if position and position != (person.position or ""):
                    person.position = position
                    changed.append(f"职位->{position}")
                if phone and phone != (person.phone or ""):
                    person.phone = phone
                    changed.append(f"电话->{phone}")
                # 范式: 单位以「项目清单」为准 —— 已有单位的人员不覆盖, 仅无单位的新增人员补单位
                if company_id and person.company_id is None:
                    person.company_id = company_id
                    changed.append(f"单位->{company_id}")
                elif company_id and person.company_id != company_id:
                    log.append(f"[{name}] 已有单位(项目清单), 保持单位->{person.company_id}, 不覆盖")
                if ext:
                    attrs = dict(person.ext_attrs or {})
                    ext_changed = []
                    for k, v in ext.items():
                        if attrs.get(k) != v:
                            attrs[k] = v
                            ext_changed.append(f"{k}->{v}")
                    if ext_changed:
                        person.ext_attrs = attrs
                        changed.extend(ext_changed)
                if changed:
                    updated += 1
                    log.append(f"[{idx}] 更新[{name}] {'; '.join(changed)}")
                else:
                    skipped += 1
                    log.append(f"[{idx}] 跳过[{name}] 字段无变化")
        except Exception as e:  # noqa: BLE001
            db.rollback()
            errors.append(f"[{idx}] {name} 导入失败: {e}")
            log.append(f"[{idx}] {name} 导入失败: {e}")
        # 实时进度(后台任务轮询)
        if progress:
            progress(stage="", imported=imported, updated=updated, skipped=skipped,
                     failed=len(errors), log=log[-1] if log else "")

    db.commit()

    # ---- Neo4j 同步: 新建/更新的人员节点 + 单位同事关系 ----
    try:
        # 重新拉取本次涉及的所有人员(按姓名), 同步节点
        involved_names = []
        for row in rows:
            nm = _cell(row, "姓名")
            if nm and nm not in involved_names:
                involved_names.append(nm)
        company_ids: set = set()
        for nm in involved_names:
            p = _find_person(db, nm)
            if not p:
                continue
            cname = ""
            cprov, ccity = "", ""
            if p.company_id:
                comp = db.get(Company, p.company_id)
                cname = comp.name if comp else ""
                cprov, ccity = (comp.province or "", comp.city or "") if comp else ("", "")
            try:
                sync_person(
                    person_id=p.id, name=p.name or "", position=p.position or "",
                    status=p.status or "active", company_id=p.company_id,
                    company_name=cname, email=p.email or "", phone=p.phone or "",
                    is_active=bool(p.is_active), province=cprov, city=ccity,
                )
            except Exception as e:  # noqa: BLE001
                errors.append(f"[{p.name}] Neo4j 同步失败: {e}")
            if p.company_id:
                company_ids.add(p.company_id)
        for cid in company_ids:
            persons = db.execute(
                select(Person).where(Person.company_id == cid, Person.is_deleted == False)
            ).scalars().all()
            if len(persons) >= 2:
                try:
                    sync_company_colleagues(cid, [
                        {"person_id": p.id, "name": p.name} for p in persons
                    ])
                except Exception:  # noqa: BLE001, S110
                    pass
        log.append(f"Neo4j 图谱同步完成({len(involved_names)} 人)")
    except Exception as e:  # noqa: BLE001
        errors.append(f"Neo4j 同步失败: {e}")

    return {
        "success": True,
        "message": f"导入完成: 新建 {imported} 人, 更新 {updated} 人, 跳过 {skipped} 人",
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "log": log,
        "errors": errors,
    }
