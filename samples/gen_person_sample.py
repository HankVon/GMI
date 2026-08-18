"""生成人员导入示例 Excel"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "persons_import"

headers = ["code", "name", "position", "status", "email", "phone", "entry_date"]
f = Font(bold=True, color="FFFFFF", size=11)
fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
for i, h in enumerate(headers, 1):
    c = ws.cell(1, i, h)
    c.font = f; c.fill = fill

rows = [
    ["EMP-001", "张三", "高级地质工程师", "active", "zhangsan@example.com", "13800001001", "2023-03-15"],
    ["EMP-002", "李四", "项目经理",       "active", "lisi@example.com",     "13800001002", "2022-01-10"],
    ["EMP-003", "王五", "勘探技术员",     "active", "wangwu@example.com",   "13800001003", "2024-06-01"],
    ["EMP-004", "赵六", "安全总监",       "active", "zhaoliu@example.com",  "13800001004", "2021-09-20"],
    ["EMP-005", "孙七", "会计",           "active", "sunqi@example.com",    "13800001005", "2023-11-01"],
]
for ri, row in enumerate(rows, 2):
    for ci, val in enumerate(row, 1):
        ws.cell(ri, ci, val)

for i in range(1, 8):
    ws.column_dimensions[ws.cell(1, i).column_letter].width = 22

path = "d:/Geology/SSM/samples/person_import_sample.xlsx"
wb.save(path)
print(f"Saved: {path}")
