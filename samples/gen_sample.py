"""生成项目导入示例 Excel"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "projects_import"

# 表头行（蓝色底白字）
headers = ["code", "name", "status", "start_date", "end_date", "description"]
f_head = Font(bold=True, color="FFFFFF", size=11)
fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
for i, h in enumerate(headers, 1):
    c = ws.cell(1, i, h)
    c.font = f_head
    c.fill = fill

# 示例数据
rows = [
    ["PRJ-001", "四川铜矿勘探项目", "active",    "2025-03-01", "2026-12-31", "四川省西部铜矿资源勘查"],
    ["PRJ-002", "云南金矿详查项目", "active",    "2025-06-15", "2027-06-30", "云南某大型金矿详查"],
    ["PRJ-003", "贵州煤矿技改项目", "completed", "2024-01-10", "2025-12-20", "煤矿安全技术改造工程"],
    ["PRJ-004", "西藏铁矿预查项目", "suspended", "2025-09-01", "2026-09-30", "高原地区铁矿预查"],
    ["PRJ-005", "湖北磷矿开发项目", "active",    "2025-01-20", "2028-01-20", ""],
]
for ri, row in enumerate(rows, 2):
    for ci, val in enumerate(row, 1):
        ws.cell(ri, ci, val)

# 列宽
for i in range(1, 7):
    ws.column_dimensions[ws.cell(1, i).column_letter].width = 20

path = "d:/Geology/SSM/samples/project_import_sample.xlsx"
wb.save(path)
print(f"Sample saved to: {path}")
